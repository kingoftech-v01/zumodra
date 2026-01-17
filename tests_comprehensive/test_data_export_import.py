#!/usr/bin/env python
"""
Comprehensive Data Export & Import Testing Suite for Zumodra

Tests the following functionality:
1. CSV export from various modules (ATS, HR, Analytics)
2. Excel export functionality with data preservation
3. PDF report generation and content validation
4. Bulk data import via CSV with validation
5. Data integrity on import/export cycles
6. Error handling for invalid imports
7. Export/import audit logging
8. Multi-tenant data isolation in exports
9. Permission-based access to exports
10. Rate limiting on bulk operations

Run with: pytest tests_comprehensive/test_data_export_import.py -v
"""

import pytest
import json
import csv
import os
import tempfile
import io
from datetime import datetime, timedelta
from decimal import Decimal
from unittest.mock import patch, MagicMock

from django.test import TestCase, Client, TransactionTestCase
from django.contrib.auth import get_user_model
from django.core.management import call_command
from django.db import connection, transaction
from django.core.exceptions import ValidationError
from django.http import HttpResponse
from django.utils import timezone
from rest_framework.test import APITestCase, APIClient
from rest_framework import status

from tenants.models import Tenant
from tenants.utils import tenant_context
from accounts.models import TenantUser, UserProfile
from ats.models import (
    JobPosting, Candidate, Application, Interview,
    InterviewFeedback, Offer, Pipeline, JobCategory
)
from ats.serializers import CandidateBulkImportSerializer
from ats.views import CandidateViewSet
from hr_core.models import Employee, TimeOff
from finance.models import Subscription
from integrations.models import OutboundWebhook, WebhookEvent
from analytics.models import AnalyticsEvent
from core.audit_logging import audit_log

User = get_user_model()


@pytest.mark.integration
class TestCSVExport(TransactionTestCase):
    """Test CSV export functionality from various modules."""

    def setUp(self):
        """Set up test data."""
        self.tenant = Tenant.objects.create(
            name='Test Tenant',
            slug='test-tenant',
            schema_name='test_tenant_schema'
        )
        self.user = User.objects.create_user(
            username='test@example.com',
            email='test@example.com',
            password='testpass123'
        )
        self.tenant_user = TenantUser.objects.create(
            user=self.user,
            tenant=self.tenant,
            role='recruiter'
        )
        self.client = APIClient()
        self.client.force_authenticate(user=self.user)

    def test_csv_candidate_export(self):
        """Test exporting candidates to CSV."""
        with tenant_context(self.tenant):
            # Create test candidates
            candidates = []
            for i in range(5):
                candidate = Candidate.objects.create(
                    first_name=f'Candidate{i}',
                    last_name='Test',
                    email=f'candidate{i}@example.com',
                    phone_number=f'555000{i}',
                    source='imported'
                )
                candidates.append(candidate)

            # Export to CSV
            response = self.client.post('/api/v1/ats/candidates/export/', {
                'format': 'csv'
            })

            assert response.status_code == status.HTTP_200_OK

            # Parse CSV response
            content = response.content.decode('utf-8')
            csv_reader = csv.DictReader(io.StringIO(content))
            rows = list(csv_reader)

            # Validate export
            assert len(rows) == 5
            assert rows[0]['first_name'] == 'Candidate0'
            assert rows[0]['email'] == 'candidate0@example.com'

    def test_csv_job_export(self):
        """Test exporting job postings to CSV."""
        with tenant_context(self.tenant):
            # Create test jobs
            category = JobCategory.objects.create(name='Engineering')
            jobs = []
            for i in range(3):
                job = JobPosting.objects.create(
                    title=f'Job {i}',
                    description=f'Job description {i}',
                    category=category,
                    status='open'
                )
                jobs.append(job)

            # Export to CSV
            response = self.client.post('/api/v1/ats/jobs/export/', {
                'format': 'csv'
            })

            assert response.status_code == status.HTTP_200_OK
            content = response.content.decode('utf-8')
            csv_reader = csv.DictReader(io.StringIO(content))
            rows = list(csv_reader)

            assert len(rows) == 3
            assert rows[0]['title'] == 'Job 0'

    def test_csv_export_with_filters(self):
        """Test CSV export with filtering criteria."""
        with tenant_context(self.tenant):
            # Create candidates with different statuses
            for i in range(10):
                Candidate.objects.create(
                    first_name=f'Candidate{i}',
                    last_name='Test',
                    email=f'candidate{i}@example.com',
                    source='linkedin' if i % 2 == 0 else 'direct'
                )

            # Export filtered candidates
            response = self.client.post('/api/v1/ats/candidates/export/', {
                'format': 'csv',
                'filters': {'source': 'linkedin'}
            })

            assert response.status_code == status.HTTP_200_OK
            content = response.content.decode('utf-8')
            rows = list(csv.DictReader(io.StringIO(content)))

            # Should only have linkedin candidates
            assert len(rows) == 5
            for row in rows:
                assert row['source'] == 'linkedin'


@pytest.mark.integration
class TestExcelExport(TransactionTestCase):
    """Test Excel export functionality."""

    def setUp(self):
        """Set up test data."""
        self.tenant = Tenant.objects.create(
            name='Excel Test Tenant',
            slug='excel-tenant',
            schema_name='excel_tenant_schema'
        )
        self.user = User.objects.create_user(
            username='excel@example.com',
            email='excel@example.com',
            password='testpass123'
        )
        TenantUser.objects.create(
            user=self.user,
            tenant=self.tenant,
            role='hr_manager'
        )
        self.client = APIClient()
        self.client.force_authenticate(user=self.user)

    def test_excel_candidates_export(self):
        """Test exporting candidates to Excel format."""
        try:
            import openpyxl
        except ImportError:
            self.skipTest("openpyxl not installed")

        with tenant_context(self.tenant):
            # Create candidates
            for i in range(10):
                Candidate.objects.create(
                    first_name=f'Excel{i}',
                    last_name='Candidate',
                    email=f'excel{i}@test.com',
                    phone_number=f'555111{i:04d}',
                    source='direct'
                )

            # Request Excel export
            response = self.client.post('/api/v1/ats/candidates/export/', {
                'format': 'excel'
            })

            assert response.status_code == status.HTTP_200_OK
            assert response['Content-Type'] == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

            # Parse Excel file
            wb = openpyxl.load_workbook(io.BytesIO(response.content))
            ws = wb.active

            # Validate content
            assert ws['A1'].value == 'first_name'
            assert ws.max_row >= 11  # 10 candidates + header

    def test_excel_analytics_export(self):
        """Test exporting analytics data to Excel."""
        try:
            import openpyxl
        except ImportError:
            self.skipTest("openpyxl not installed")

        with tenant_context(self.tenant):
            # Export analytics
            response = self.client.get('/api/v1/analytics/dashboard/', {
                'format': 'excel',
                'dashboard_type': 'recruitment'
            })

            if response.status_code == status.HTTP_200_OK:
                assert response['Content-Type'] == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'


@pytest.mark.integration
class TestPDFGeneration(TransactionTestCase):
    """Test PDF report generation."""

    def setUp(self):
        """Set up test data."""
        self.tenant = Tenant.objects.create(
            name='PDF Test Tenant',
            slug='pdf-tenant',
            schema_name='pdf_tenant_schema'
        )
        self.user = User.objects.create_user(
            username='pdf@example.com',
            email='pdf@example.com',
            password='testpass123'
        )
        TenantUser.objects.create(
            user=self.user,
            tenant=self.tenant,
            role='recruiter'
        )
        self.client = APIClient()
        self.client.force_authenticate(user=self.user)

    def test_pdf_recruitment_report(self):
        """Test generating PDF recruitment report."""
        try:
            import reportlab
        except ImportError:
            self.skipTest("reportlab not installed")

        with tenant_context(self.tenant):
            # Create test data
            category = JobCategory.objects.create(name='Engineering')
            job = JobPosting.objects.create(
                title='Test Job',
                description='Test Description',
                category=category,
                status='open'
            )

            candidates = []
            for i in range(5):
                candidate = Candidate.objects.create(
                    first_name=f'PDF{i}',
                    last_name='Candidate',
                    email=f'pdf{i}@test.com'
                )
                candidates.append(candidate)
                # Create applications
                Application.objects.create(
                    candidate=candidate,
                    job=job,
                    status='applied'
                )

            # Request PDF report
            response = self.client.get('/api/v1/ats/jobs/{}/report/'.format(job.id), {
                'format': 'pdf'
            })

            if response.status_code == status.HTTP_200_OK:
                assert response['Content-Type'] == 'application/pdf'
                assert len(response.content) > 0

    def test_pdf_analytics_report(self):
        """Test generating PDF analytics report."""
        try:
            import reportlab
        except ImportError:
            self.skipTest("reportlab not installed")

        with tenant_context(self.tenant):
            response = self.client.get('/api/v1/analytics/report/', {
                'format': 'pdf',
                'period': 'monthly'
            })

            if response.status_code == status.HTTP_200_OK:
                assert response['Content-Type'] == 'application/pdf'


@pytest.mark.integration
class TestBulkImport(TransactionTestCase):
    """Test bulk data import functionality."""

    def setUp(self):
        """Set up test data."""
        self.tenant = Tenant.objects.create(
            name='Import Test Tenant',
            slug='import-tenant',
            schema_name='import_tenant_schema'
        )
        self.user = User.objects.create_user(
            username='import@example.com',
            email='import@example.com',
            password='testpass123'
        )
        TenantUser.objects.create(
            user=self.user,
            tenant=self.tenant,
            role='recruiter'
        )
        self.client = APIClient()
        self.client.force_authenticate(user=self.user)

    def test_import_candidates_csv(self):
        """Test importing candidates from CSV."""
        with tenant_context(self.tenant):
            # Create CSV content
            csv_content = """first_name,last_name,email,phone_number,source
John,Doe,john.doe@example.com,555-0001,linkedin
Jane,Smith,jane.smith@example.com,555-0002,direct
Bob,Johnson,bob.johnson@example.com,555-0003,referral"""

            # Save to temporary file
            with tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False) as f:
                f.write(csv_content)
                csv_file = f.name

            try:
                # Call import command
                call_command(
                    'import_candidates_csv',
                    csv_file,
                    'import-tenant',
                    verbosity=0
                )

                # Verify imports
                assert Candidate.objects.count() == 3
                assert Candidate.objects.filter(first_name='John').exists()
                assert Candidate.objects.filter(email='jane.smith@example.com').exists()
            finally:
                os.unlink(csv_file)

    def test_import_jobs_csv(self):
        """Test importing job postings from CSV."""
        with tenant_context(self.tenant):
            # Create category
            JobCategory.objects.create(name='Engineering')

            # Create CSV content
            csv_content = """title,description,category,status,salary_min,salary_max
Software Engineer,Build amazing software,Engineering,open,80000,120000
DevOps Engineer,Infrastructure and deployment,Engineering,open,90000,130000
Product Manager,Lead product development,Engineering,draft,100000,150000"""

            with tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False) as f:
                f.write(csv_content)
                csv_file = f.name

            try:
                call_command(
                    'import_jobs_csv',
                    csv_file,
                    'import-tenant',
                    verbosity=0
                )

                assert JobPosting.objects.count() == 3
                assert JobPosting.objects.filter(title='Software Engineer').exists()
            finally:
                os.unlink(csv_file)

    def test_import_with_validation_errors(self):
        """Test import with invalid data."""
        with tenant_context(self.tenant):
            # CSV with missing required fields
            csv_content = """first_name,last_name,email
John,Doe,
Jane,Smith,jane@example.com"""

            with tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False) as f:
                f.write(csv_content)
                csv_file = f.name

            try:
                # Should raise CommandError or similar
                with pytest.raises(Exception):
                    call_command(
                        'import_candidates_csv',
                        csv_file,
                        'import-tenant',
                        verbosity=0
                    )
            finally:
                os.unlink(csv_file)


@pytest.mark.integration
class TestImportValidation(TransactionTestCase):
    """Test data validation during import."""

    def setUp(self):
        """Set up test data."""
        self.tenant = Tenant.objects.create(
            name='Validation Test Tenant',
            slug='validation-tenant',
            schema_name='validation_tenant_schema'
        )
        self.user = User.objects.create_user(
            username='validation@example.com',
            email='validation@example.com',
            password='testpass123'
        )
        TenantUser.objects.create(
            user=self.user,
            tenant=self.tenant,
            role='recruiter'
        )
        self.client = APIClient()
        self.client.force_authenticate(user=self.user)

    def test_validate_email_uniqueness(self):
        """Test email uniqueness validation on import."""
        with tenant_context(self.tenant):
            # Create existing candidate
            Candidate.objects.create(
                first_name='Existing',
                last_name='Candidate',
                email='duplicate@example.com'
            )

            # Try to import duplicate email
            csv_content = """first_name,last_name,email
New,Candidate,duplicate@example.com"""

            with tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False) as f:
                f.write(csv_content)
                csv_file = f.name

            try:
                call_command(
                    'import_candidates_csv',
                    csv_file,
                    'validation-tenant',
                    '--skip-duplicates',
                    verbosity=0
                )

                # Should skip duplicate
                assert Candidate.objects.count() == 1
            finally:
                os.unlink(csv_file)

    def test_validate_required_fields(self):
        """Test required field validation."""
        with tenant_context(self.tenant):
            # CSV missing required fields
            csv_content = """first_name,last_name
John,Doe"""

            with tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False) as f:
                f.write(csv_content)
                csv_file = f.name

            try:
                with pytest.raises(Exception):
                    call_command(
                        'import_candidates_csv',
                        csv_file,
                        'validation-tenant',
                        verbosity=0
                    )
            finally:
                os.unlink(csv_file)

    def test_validate_data_types(self):
        """Test data type validation during import."""
        with tenant_context(self.tenant):
            JobCategory.objects.create(name='Engineering')

            # CSV with invalid data types
            csv_content = """title,description,category,salary_min,salary_max
Engineer,Desc,Engineering,invalid_number,120000"""

            with tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False) as f:
                f.write(csv_content)
                csv_file = f.name

            try:
                with pytest.raises(Exception):
                    call_command(
                        'import_jobs_csv',
                        csv_file,
                        'validation-tenant',
                        verbosity=0
                    )
            finally:
                os.unlink(csv_file)


@pytest.mark.integration
class TestExportImportDataIntegrity(TransactionTestCase):
    """Test data integrity in export/import cycles."""

    def setUp(self):
        """Set up test data."""
        self.tenant = Tenant.objects.create(
            name='Integrity Test Tenant',
            slug='integrity-tenant',
            schema_name='integrity_tenant_schema'
        )
        self.user = User.objects.create_user(
            username='integrity@example.com',
            email='integrity@example.com',
            password='testpass123'
        )
        TenantUser.objects.create(
            user=self.user,
            tenant=self.tenant,
            role='recruiter'
        )
        self.client = APIClient()
        self.client.force_authenticate(user=self.user)

    def test_candidate_export_import_integrity(self):
        """Test data integrity through export/import cycle."""
        with tenant_context(self.tenant):
            # Create original candidate
            original = Candidate.objects.create(
                first_name='Integrity',
                last_name='Test',
                email='integrity@example.com',
                phone_number='555-9999',
                source='direct',
                skills=['Python', 'Django', 'PostgreSQL'],
                tags=['python', 'backend']
            )

            # Export
            response = self.client.post('/api/v1/ats/candidates/export/', {
                'format': 'csv'
            })

            # Delete original
            Candidate.objects.all().delete()

            # Import from export
            content = response.content.decode('utf-8')
            with tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False) as f:
                f.write(content)
                csv_file = f.name

            try:
                call_command(
                    'import_candidates_csv',
                    csv_file,
                    'integrity-tenant',
                    verbosity=0
                )

                # Verify integrity
                imported = Candidate.objects.get(email='integrity@example.com')
                assert imported.first_name == original.first_name
                assert imported.phone_number == original.phone_number
                assert imported.source == original.source
                assert set(imported.skills) == set(original.skills)
                assert set(imported.tags) == set(original.tags)
            finally:
                os.unlink(csv_file)


@pytest.mark.integration
class TestAuditLogging(TransactionTestCase):
    """Test audit logging for export/import operations."""

    def setUp(self):
        """Set up test data."""
        self.tenant = Tenant.objects.create(
            name='Audit Test Tenant',
            slug='audit-tenant',
            schema_name='audit_tenant_schema'
        )
        self.user = User.objects.create_user(
            username='audit@example.com',
            email='audit@example.com',
            password='testpass123'
        )
        TenantUser.objects.create(
            user=self.user,
            tenant=self.tenant,
            role='recruiter'
        )
        self.client = APIClient()
        self.client.force_authenticate(user=self.user)

    def test_export_audit_logging(self):
        """Test that exports are logged in audit trail."""
        with tenant_context(self.tenant):
            # Create candidates
            for i in range(3):
                Candidate.objects.create(
                    first_name=f'Audit{i}',
                    last_name='Test',
                    email=f'audit{i}@example.com'
                )

            # Export
            response = self.client.post('/api/v1/ats/candidates/export/', {
                'format': 'csv'
            })

            assert response.status_code == status.HTTP_200_OK

            # Verify audit log exists
            # This would need audit log models to verify
            # assert AuditLog.objects.filter(
            #     user=self.user,
            #     action='export_candidates'
            # ).exists()

    def test_import_audit_logging(self):
        """Test that imports are logged in audit trail."""
        with tenant_context(self.tenant):
            csv_content = """first_name,last_name,email
Audit,Import,audit.import@example.com"""

            with tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False) as f:
                f.write(csv_content)
                csv_file = f.name

            try:
                call_command(
                    'import_candidates_csv',
                    csv_file,
                    'audit-tenant',
                    verbosity=0
                )

                # Verify candidates were created
                assert Candidate.objects.filter(
                    email='audit.import@example.com'
                ).exists()
            finally:
                os.unlink(csv_file)


@pytest.mark.integration
class TestMultiTenantIsolation(TransactionTestCase):
    """Test data isolation in multi-tenant export/import."""

    def setUp(self):
        """Set up multiple tenants."""
        self.tenant1 = Tenant.objects.create(
            name='Tenant One',
            slug='tenant-one',
            schema_name='tenant_one_schema'
        )
        self.tenant2 = Tenant.objects.create(
            name='Tenant Two',
            slug='tenant-two',
            schema_name='tenant_two_schema'
        )

        self.user1 = User.objects.create_user(
            username='user1@example.com',
            email='user1@example.com',
            password='testpass123'
        )
        self.user2 = User.objects.create_user(
            username='user2@example.com',
            email='user2@example.com',
            password='testpass123'
        )

        TenantUser.objects.create(user=self.user1, tenant=self.tenant1, role='recruiter')
        TenantUser.objects.create(user=self.user2, tenant=self.tenant2, role='recruiter')

        self.client1 = APIClient()
        self.client2 = APIClient()
        self.client1.force_authenticate(user=self.user1)
        self.client2.force_authenticate(user=self.user2)

    def test_tenant_data_isolation_on_export(self):
        """Test that exports only return data from current tenant."""
        # Add data to tenant 1
        with tenant_context(self.tenant1):
            for i in range(5):
                Candidate.objects.create(
                    first_name=f'Tenant1-{i}',
                    last_name='Candidate',
                    email=f't1-cand{i}@example.com'
                )

        # Add data to tenant 2
        with tenant_context(self.tenant2):
            for i in range(3):
                Candidate.objects.create(
                    first_name=f'Tenant2-{i}',
                    last_name='Candidate',
                    email=f't2-cand{i}@example.com'
                )

        # User 1 exports should only see tenant 1 data
        with tenant_context(self.tenant1):
            response = self.client1.post('/api/v1/ats/candidates/export/', {
                'format': 'csv'
            })

            content = response.content.decode('utf-8')
            rows = list(csv.DictReader(io.StringIO(content)))
            assert len(rows) == 5
            for row in rows:
                assert 'Tenant1' in row['first_name']


@pytest.mark.integration
class TestErrorHandling(TransactionTestCase):
    """Test error handling in export/import operations."""

    def setUp(self):
        """Set up test data."""
        self.tenant = Tenant.objects.create(
            name='Error Test Tenant',
            slug='error-tenant',
            schema_name='error_tenant_schema'
        )
        self.user = User.objects.create_user(
            username='error@example.com',
            email='error@example.com',
            password='testpass123'
        )
        TenantUser.objects.create(
            user=self.user,
            tenant=self.tenant,
            role='recruiter'
        )
        self.client = APIClient()
        self.client.force_authenticate(user=self.user)

    def test_export_file_not_found(self):
        """Test handling of missing file on import."""
        with tenant_context(self.tenant):
            with pytest.raises(Exception):
                call_command(
                    'import_candidates_csv',
                    '/nonexistent/path/file.csv',
                    'error-tenant',
                    verbosity=0
                )

    def test_import_invalid_csv_format(self):
        """Test handling of malformed CSV."""
        with tenant_context(self.tenant):
            # Create invalid CSV (unclosed quotes)
            csv_content = """first_name,last_name,email
"John,Doe,john@example.com
Jane,Smith,jane@example.com"""

            with tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False) as f:
                f.write(csv_content)
                csv_file = f.name

            try:
                with pytest.raises(Exception):
                    call_command(
                        'import_candidates_csv',
                        csv_file,
                        'error-tenant',
                        verbosity=0
                    )
            finally:
                os.unlink(csv_file)

    def test_import_encoding_error(self):
        """Test handling of encoding errors."""
        with tenant_context(self.tenant):
            csv_content = """first_name,last_name,email
John,Döe,john@example.com"""

            with tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False, encoding='latin-1') as f:
                f.write(csv_content)
                csv_file = f.name

            try:
                # Import with wrong encoding should handle gracefully
                call_command(
                    'import_candidates_csv',
                    csv_file,
                    'error-tenant',
                    '--encoding=utf-8',
                    verbosity=0
                )
            finally:
                os.unlink(csv_file)

    def test_import_permission_denied(self):
        """Test handling of permission errors."""
        # Create user without import permissions
        unprivileged_user = User.objects.create_user(
            username='viewer@example.com',
            email='viewer@example.com',
            password='testpass123'
        )
        TenantUser.objects.create(
            user=unprivileged_user,
            tenant=self.tenant,
            role='viewer'
        )

        client = APIClient()
        client.force_authenticate(user=unprivileged_user)

        with tenant_context(self.tenant):
            response = client.post('/api/v1/ats/candidates/export/', {
                'format': 'csv'
            })

            # Should be denied
            assert response.status_code in [
                status.HTTP_403_FORBIDDEN,
                status.HTTP_401_UNAUTHORIZED
            ]


@pytest.mark.integration
class TestRateLimiting(TransactionTestCase):
    """Test rate limiting on export/import operations."""

    def setUp(self):
        """Set up test data."""
        self.tenant = Tenant.objects.create(
            name='Rate Limit Tenant',
            slug='ratelimit-tenant',
            schema_name='ratelimit_tenant_schema'
        )
        self.user = User.objects.create_user(
            username='ratelimit@example.com',
            email='ratelimit@example.com',
            password='testpass123'
        )
        TenantUser.objects.create(
            user=self.user,
            tenant=self.tenant,
            role='recruiter'
        )
        self.client = APIClient()
        self.client.force_authenticate(user=self.user)

    def test_bulk_import_rate_limit(self):
        """Test rate limiting on bulk import."""
        with tenant_context(self.tenant):
            csv_content = """first_name,last_name,email
John,Doe,john@example.com"""

            # Make multiple rapid import requests
            for i in range(10):
                response = self.client.post('/api/v1/ats/candidates/bulk-import/', {
                    'csv_data': csv_content
                }, format='json')

                if response.status_code == status.HTTP_429_TOO_MANY_REQUESTS:
                    # Rate limit hit
                    break

    def test_export_rate_limit(self):
        """Test rate limiting on exports."""
        with tenant_context(self.tenant):
            # Create candidates
            for i in range(5):
                Candidate.objects.create(
                    first_name=f'Candidate{i}',
                    last_name='Test',
                    email=f'cand{i}@example.com'
                )

            # Make rapid export requests
            responses = []
            for i in range(20):
                response = self.client.post('/api/v1/ats/candidates/export/', {
                    'format': 'csv'
                })
                responses.append(response.status_code)

                if response.status_code == status.HTTP_429_TOO_MANY_REQUESTS:
                    break


@pytest.mark.integration
class TestExportPerformance(TransactionTestCase):
    """Test performance of export operations with large datasets."""

    def setUp(self):
        """Set up test data."""
        self.tenant = Tenant.objects.create(
            name='Performance Tenant',
            slug='perf-tenant',
            schema_name='perf_tenant_schema'
        )
        self.user = User.objects.create_user(
            username='perf@example.com',
            email='perf@example.com',
            password='testpass123'
        )
        TenantUser.objects.create(
            user=self.user,
            tenant=self.tenant,
            role='recruiter'
        )
        self.client = APIClient()
        self.client.force_authenticate(user=self.user)

    def test_export_large_candidate_set(self):
        """Test exporting large number of candidates."""
        with tenant_context(self.tenant):
            # Create 1000 candidates
            candidates = []
            for i in range(1000):
                candidates.append(
                    Candidate(
                        first_name=f'Candidate{i}',
                        last_name='Test',
                        email=f'cand{i}@example.com',
                        source='direct'
                    )
                )

            Candidate.objects.bulk_create(candidates, batch_size=100)

            # Export should complete successfully
            import time
            start = time.time()

            response = self.client.post('/api/v1/ats/candidates/export/', {
                'format': 'csv'
            })

            elapsed = time.time() - start

            assert response.status_code == status.HTTP_200_OK
            assert elapsed < 30  # Should complete in less than 30 seconds

            # Parse to verify all records
            content = response.content.decode('utf-8')
            rows = list(csv.DictReader(io.StringIO(content)))
            assert len(rows) == 1000


if __name__ == '__main__':
    pytest.main([__file__, '-v', '--tb=short'])

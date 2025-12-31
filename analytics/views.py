"""
Analytics Views - Dashboard API Endpoints

This module provides API views for:
- RecruitmentDashboardView: Recruitment metrics and KPIs
- DiversityDashboardView: Anonymized diversity analytics
- HRDashboardView: HR metrics (retention, time-off, performance)
- ExecutiveSummaryView: High-level executive summary
- ExportReportView: PDF/Excel report exports
"""

import io
from datetime import date, datetime, timedelta
from typing import Optional

from django.shortcuts import render
from django.http import HttpResponse, FileResponse
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.utils import timezone
from django.db.models import Count, Sum, Avg, Q
from django.db.models.functions import TruncDate, TruncMonth

from rest_framework import status, generics
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework.decorators import api_view, permission_classes

from .models import (
    PageView, UserAction, SearchQuery, DashboardMetric,
    RecruitmentMetric, DiversityMetric, HiringFunnelMetric,
    TimeToHireMetric, SourceEffectivenessMetric, EmployeeRetentionMetric,
    TimeOffAnalytics, PerformanceDistribution, DashboardCache
)
from .serializers import (
    DateRangeSerializer, RecruitmentDashboardSerializer,
    HRDashboardSerializer, ExecutiveSummarySerializer,
    AnonymizedDiversitySerializer, RecruitmentMetricSerializer,
    DiversityMetricSerializer, HiringFunnelMetricSerializer,
    TimeToHireMetricSerializer, SourceEffectivenessMetricSerializer,
    EmployeeRetentionMetricSerializer, TimeOffAnalyticsSerializer,
    PerformanceDistributionSerializer, ExportRequestSerializer,
    ExportResponseSerializer, FunnelStageSerializer
)
from .services import (
    DateRangeFilter, RecruitmentAnalyticsService, DiversityAnalyticsService,
    HRAnalyticsService, DashboardDataService
)


# ==================== TEMPLATE-BASED VIEWS (Existing) ====================

@login_required
def analytics_dashboard(request):
    """
    Main analytics dashboard with key metrics (template-based).
    """
    today = timezone.now().date()
    last_7_days = today - timedelta(days=7)
    last_30_days = today - timedelta(days=30)

    # Lazy imports to avoid circular imports
    try:
        from custom_account_u.models import CustomUser as User
        from services.models import DService, DServiceContract, DServiceProviderProfile

        # User metrics
        total_users = User.objects.count()
        active_users_7d = User.objects.filter(last_login__gte=last_7_days).count()
        new_users_7d = User.objects.filter(date_joined__gte=last_7_days).count()

        # Service metrics
        total_services = DService.objects.count()
        new_services_7d = DService.objects.filter(created_at__gte=last_7_days).count()

        # Contract metrics
        total_contracts = DServiceContract.objects.count()
        active_contracts = DServiceContract.objects.filter(status='active').count()
        completed_contracts = DServiceContract.objects.filter(status='completed').count()

        # Provider metrics
        total_providers = DServiceProviderProfile.objects.count()
        verified_providers = DServiceProviderProfile.objects.filter(is_verified=True).count()
        avg_provider_rating = DServiceProviderProfile.objects.aggregate(
            avg=Avg('rating_avg')
        )['avg'] or 0

        # Popular services
        popular_services = DService.objects.annotate(
            likes_count=Count('config_liked_DServices')
        ).order_by('-likes_count')[:5]

    except Exception:
        # Fallback if models don't exist
        total_users = 0
        active_users_7d = 0
        new_users_7d = 0
        total_services = 0
        new_services_7d = 0
        total_contracts = 0
        active_contracts = 0
        completed_contracts = 0
        total_providers = 0
        verified_providers = 0
        avg_provider_rating = 0
        popular_services = []

    # Page views
    page_views_7d = PageView.objects.filter(timestamp__gte=last_7_days).count()

    # Recent user actions
    recent_actions = UserAction.objects.select_related('user')[:20]

    # Popular searches
    popular_searches = SearchQuery.objects.filter(
        timestamp__gte=last_7_days
    ).values('query').annotate(
        count=Count('id')
    ).order_by('-count')[:10]

    context = {
        'total_users': total_users,
        'active_users_7d': active_users_7d,
        'new_users_7d': new_users_7d,
        'total_services': total_services,
        'new_services_7d': new_services_7d,
        'popular_services': popular_services,
        'total_contracts': total_contracts,
        'active_contracts': active_contracts,
        'completed_contracts': completed_contracts,
        'total_providers': total_providers,
        'verified_providers': verified_providers,
        'avg_provider_rating': round(avg_provider_rating, 2) if avg_provider_rating else 0,
        'page_views_7d': page_views_7d,
        'recent_actions': recent_actions,
        'popular_searches': popular_searches,
    }

    return render(request, 'analytics/dashboard.html', context)


@login_required
def provider_analytics(request):
    """Analytics dashboard for service providers (template-based)."""
    try:
        from services.models import DServiceProviderProfile
        provider = request.user.DService_provider_profile
    except Exception:
        return render(request, 'analytics/no_provider_profile.html')

    context = {
        'provider': provider,
    }

    return render(request, 'analytics/provider_analytics.html', context)


@login_required
def client_analytics(request):
    """Analytics dashboard for clients (template-based)."""
    context = {}
    return render(request, 'analytics/client_analytics.html', context)


# ==================== API VIEWS ====================

class BaseAnalyticsAPIView(APIView):
    """Base class for analytics API views."""
    permission_classes = [IsAuthenticated]

    def get_date_filter(self, request) -> DateRangeFilter:
        """Parse and validate date range from request."""
        serializer = DateRangeSerializer(data=request.query_params)
        serializer.is_valid(raise_exception=True)
        data = serializer.validated_data

        return DateRangeFilter(
            start_date=data.get('start_date'),
            end_date=data.get('end_date'),
            period=data.get('period', 'month')
        )


class RecruitmentDashboardView(BaseAnalyticsAPIView):
    """
    API endpoint for recruitment dashboard data.

    GET /api/analytics/recruitment/

    Query Parameters:
    - start_date: Start date (YYYY-MM-DD)
    - end_date: End date (YYYY-MM-DD)
    - period: day|week|month|quarter|year (default: month)

    Returns recruitment KPIs including:
    - Job metrics (open positions, new positions, etc.)
    - Application metrics (total, by status, by source)
    - Interview metrics (scheduled, completed, etc.)
    - Offer metrics (extended, accepted, declined)
    - Conversion rates
    - Time-to-hire statistics
    """

    def get(self, request):
        date_filter = self.get_date_filter(request)

        service = DashboardDataService(date_filter)
        data = service.get_recruitment_dashboard()

        serializer = RecruitmentDashboardSerializer(data)
        return Response(serializer.data)


class DiversityDashboardView(BaseAnalyticsAPIView):
    """
    API endpoint for diversity analytics (EEOC compliant, anonymized).

    GET /api/analytics/diversity/

    Query Parameters:
    - start_date: Start date (YYYY-MM-DD)
    - end_date: End date (YYYY-MM-DD)
    - scope: employees|applicants|hired|leadership (default: employees)

    Returns anonymized diversity metrics including:
    - Gender distribution
    - Ethnicity distribution (EEOC categories)
    - Age distribution (ranges)
    - Veteran/disability status
    - Department breakdowns (where data permits)

    Note: Categories with fewer than 5 individuals are suppressed
    for privacy/anonymization compliance.
    """

    def get(self, request):
        date_filter = self.get_date_filter(request)
        scope = request.query_params.get('scope', 'employees')

        if scope not in ['employees', 'applicants', 'hired', 'leadership', 'departed']:
            return Response(
                {'error': 'Invalid scope. Must be one of: employees, applicants, hired, leadership, departed'},
                status=status.HTTP_400_BAD_REQUEST
            )

        service = DashboardDataService(date_filter)
        data = service.get_diversity_dashboard(scope)

        serializer = AnonymizedDiversitySerializer(data)
        return Response(serializer.data)


class HRDashboardView(BaseAnalyticsAPIView):
    """
    API endpoint for HR dashboard data.

    GET /api/analytics/hr/

    Query Parameters:
    - start_date: Start date (YYYY-MM-DD)
    - end_date: End date (YYYY-MM-DD)
    - period: day|week|month|quarter|year (default: month)

    Returns HR metrics including:
    - Headcount (current, changes, by department/type)
    - Retention/turnover rates
    - Time-off analytics
    - Performance review metrics
    """

    def get(self, request):
        date_filter = self.get_date_filter(request)

        service = DashboardDataService(date_filter)
        data = service.get_hr_dashboard()

        serializer = HRDashboardSerializer(data)
        return Response(serializer.data)


class ExecutiveSummaryView(BaseAnalyticsAPIView):
    """
    API endpoint for executive summary dashboard.

    GET /api/analytics/executive/

    Query Parameters:
    - start_date: Start date (YYYY-MM-DD)
    - end_date: End date (YYYY-MM-DD)
    - period: day|week|month|quarter|year (default: month)

    Returns high-level summary including:
    - Key recruitment metrics
    - Key HR metrics
    - Chart data for visualizations
    """

    def get(self, request):
        date_filter = self.get_date_filter(request)

        service = DashboardDataService(date_filter)
        data = service.get_executive_summary()

        serializer = ExecutiveSummarySerializer(data)
        return Response(serializer.data)


class HiringFunnelView(BaseAnalyticsAPIView):
    """
    API endpoint for hiring funnel analytics.

    GET /api/analytics/funnel/

    Query Parameters:
    - start_date: Start date (YYYY-MM-DD)
    - end_date: End date (YYYY-MM-DD)
    - job_id: Optional job posting ID to filter by
    - department: Optional department filter

    Returns funnel data with conversion rates at each stage.
    """

    def get(self, request):
        date_filter = self.get_date_filter(request)
        job_id = request.query_params.get('job_id')
        department = request.query_params.get('department')

        # Get funnel metrics
        recruitment_service = RecruitmentAnalyticsService(date_filter)
        conversion_rates = recruitment_service.get_conversion_rates()
        app_metrics = recruitment_service.get_application_metrics()

        # Build funnel data
        funnel_stages = [
            {
                'stage': 'Applied',
                'count': app_metrics['total_applications'],
                'conversion_rate': 100,
                'avg_days': None,
            },
            {
                'stage': 'Screened',
                'count': app_metrics['shortlisted'],
                'conversion_rate': conversion_rates['applied_to_interview'],
                'avg_days': None,
            },
            {
                'stage': 'Interviewing',
                'count': app_metrics['interviewing'],
                'conversion_rate': None,
                'avg_days': None,
            },
            {
                'stage': 'Offered',
                'count': recruitment_service.get_offer_metrics()['total_offers'],
                'conversion_rate': conversion_rates['interview_to_offer'],
                'avg_days': None,
            },
            {
                'stage': 'Hired',
                'count': app_metrics['hired'],
                'conversion_rate': conversion_rates['offer_to_hire'],
                'avg_days': None,
            },
        ]

        serializer = FunnelStageSerializer(funnel_stages, many=True)
        return Response({
            'stages': serializer.data,
            'overall_conversion_rate': conversion_rates['overall_conversion'],
            'period': {
                'start': date_filter.start_date.isoformat(),
                'end': date_filter.end_date.isoformat(),
            },
        })


class SourceEffectivenessView(BaseAnalyticsAPIView):
    """
    API endpoint for source effectiveness analytics.

    GET /api/analytics/sources/

    Query Parameters:
    - start_date: Start date (YYYY-MM-DD)
    - end_date: End date (YYYY-MM-DD)

    Returns effectiveness metrics for each candidate source.
    """

    def get(self, request):
        date_filter = self.get_date_filter(request)

        metrics = SourceEffectivenessMetric.objects.filter(
            period_start__gte=date_filter.start_date,
            period_end__lte=date_filter.end_date,
        ).order_by('-effectiveness_score')

        serializer = SourceEffectivenessMetricSerializer(metrics, many=True)
        return Response({
            'sources': serializer.data,
            'period': {
                'start': date_filter.start_date.isoformat(),
                'end': date_filter.end_date.isoformat(),
            },
        })


class TimeToHireAnalyticsView(BaseAnalyticsAPIView):
    """
    API endpoint for time-to-hire analytics.

    GET /api/analytics/time-to-hire/

    Query Parameters:
    - start_date: Start date (YYYY-MM-DD)
    - end_date: End date (YYYY-MM-DD)
    - department: Optional department filter
    - job_type: Optional job type filter

    Returns time-to-hire statistics and trends.
    """

    def get(self, request):
        date_filter = self.get_date_filter(request)
        department = request.query_params.get('department')
        job_type = request.query_params.get('job_type')

        recruitment_service = RecruitmentAnalyticsService(date_filter)
        time_stats = recruitment_service.get_time_to_hire_stats()

        return Response({
            'statistics': time_stats,
            'filters': {
                'department': department,
                'job_type': job_type,
            },
            'period': {
                'start': date_filter.start_date.isoformat(),
                'end': date_filter.end_date.isoformat(),
            },
        })


class RetentionAnalyticsView(BaseAnalyticsAPIView):
    """
    API endpoint for employee retention analytics.

    GET /api/analytics/retention/

    Query Parameters:
    - start_date: Start date (YYYY-MM-DD)
    - end_date: End date (YYYY-MM-DD)
    - department: Optional department filter

    Returns retention and turnover metrics.
    """

    def get(self, request):
        date_filter = self.get_date_filter(request)

        hr_service = HRAnalyticsService(date_filter)
        retention_data = hr_service.get_retention_metrics()
        headcount_data = hr_service.get_headcount_metrics()

        return Response({
            'retention': retention_data,
            'headcount': headcount_data,
            'period': {
                'start': date_filter.start_date.isoformat(),
                'end': date_filter.end_date.isoformat(),
            },
        })


class PerformanceAnalyticsView(BaseAnalyticsAPIView):
    """
    API endpoint for performance review analytics.

    GET /api/analytics/performance/

    Query Parameters:
    - start_date: Start date (YYYY-MM-DD)
    - end_date: End date (YYYY-MM-DD)
    - department: Optional department filter
    - review_cycle: Optional cycle filter (annual, mid_year, etc.)

    Returns performance distribution and metrics.
    """

    def get(self, request):
        date_filter = self.get_date_filter(request)

        hr_service = HRAnalyticsService(date_filter)
        performance_data = hr_service.get_performance_metrics()

        return Response({
            'performance': performance_data,
            'period': {
                'start': date_filter.start_date.isoformat(),
                'end': date_filter.end_date.isoformat(),
            },
        })


class TimeOffAnalyticsView(BaseAnalyticsAPIView):
    """
    API endpoint for time-off analytics.

    GET /api/analytics/time-off/

    Query Parameters:
    - start_date: Start date (YYYY-MM-DD)
    - end_date: End date (YYYY-MM-DD)
    - department: Optional department filter

    Returns time-off and absence metrics.
    """

    def get(self, request):
        date_filter = self.get_date_filter(request)

        hr_service = HRAnalyticsService(date_filter)
        time_off_data = hr_service.get_time_off_metrics()

        return Response({
            'time_off': time_off_data,
            'period': {
                'start': date_filter.start_date.isoformat(),
                'end': date_filter.end_date.isoformat(),
            },
        })


class ExportReportView(BaseAnalyticsAPIView):
    """
    API endpoint for exporting analytics reports.

    POST /api/analytics/export/

    Request Body:
    - format: pdf|excel|csv (default: excel)
    - dashboard_type: recruitment|diversity|hr|executive|all
    - start_date: Start date (YYYY-MM-DD)
    - end_date: End date (YYYY-MM-DD)
    - include_charts: boolean (default: true)
    - include_comparison: boolean (default: true)

    Returns a file download or URL to the generated report.
    """

    def post(self, request):
        serializer = ExportRequestSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        data = serializer.validated_data

        export_format = data.get('format', 'excel')
        dashboard_type = data.get('dashboard_type', 'all')
        start_date = data.get('start_date', timezone.now().date() - timedelta(days=30))
        end_date = data.get('end_date', timezone.now().date())

        # Get data based on dashboard type
        date_filter = DateRangeFilter(start_date, end_date)
        service = DashboardDataService(date_filter)

        if dashboard_type == 'recruitment':
            report_data = service.get_recruitment_dashboard()
        elif dashboard_type == 'diversity':
            report_data = service.get_diversity_dashboard()
        elif dashboard_type == 'hr':
            report_data = service.get_hr_dashboard()
        elif dashboard_type == 'executive':
            report_data = service.get_executive_summary()
        else:
            report_data = {
                'recruitment': service.get_recruitment_dashboard(),
                'hr': service.get_hr_dashboard(),
                'executive': service.get_executive_summary(),
            }

        # Generate report based on format
        if export_format == 'excel':
            return self._generate_excel(report_data, dashboard_type)
        elif export_format == 'csv':
            return self._generate_csv(report_data, dashboard_type)
        elif export_format == 'pdf':
            return self._generate_pdf(report_data, dashboard_type)
        else:
            return Response(
                {'error': 'Unsupported format'},
                status=status.HTTP_400_BAD_REQUEST
            )

    def _generate_excel(self, data, dashboard_type):
        """Generate Excel report."""
        try:
            import openpyxl
            from openpyxl import Workbook
        except ImportError:
            return Response(
                {'error': 'openpyxl not installed. Install with: pip install openpyxl'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

        wb = Workbook()
        ws = wb.active
        ws.title = f"{dashboard_type.title()} Analytics"

        # Add header
        ws['A1'] = f"{dashboard_type.title()} Analytics Report"
        ws['A2'] = f"Generated: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A3'] = ""

        # Add data
        row = 5
        for key, value in data.items():
            if isinstance(value, dict):
                ws[f'A{row}'] = key.replace('_', ' ').title()
                row += 1
                for sub_key, sub_value in value.items():
                    ws[f'B{row}'] = sub_key.replace('_', ' ').title()
                    ws[f'C{row}'] = str(sub_value) if sub_value is not None else 'N/A'
                    row += 1
                row += 1
            else:
                ws[f'A{row}'] = key.replace('_', ' ').title()
                ws[f'B{row}'] = str(value) if value is not None else 'N/A'
                row += 1

        # Create response
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"{dashboard_type}_analytics_{timezone.now().strftime('%Y%m%d')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        return response

    def _generate_csv(self, data, dashboard_type):
        """Generate CSV report."""
        import csv

        buffer = io.StringIO()
        writer = csv.writer(buffer)

        # Header
        writer.writerow([f"{dashboard_type.title()} Analytics Report"])
        writer.writerow([f"Generated: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}"])
        writer.writerow([])

        # Data
        for key, value in data.items():
            if isinstance(value, dict):
                writer.writerow([key.replace('_', ' ').title()])
                for sub_key, sub_value in value.items():
                    writer.writerow(['', sub_key.replace('_', ' ').title(), str(sub_value) if sub_value is not None else 'N/A'])
                writer.writerow([])
            else:
                writer.writerow([key.replace('_', ' ').title(), str(value) if value is not None else 'N/A'])

        response = HttpResponse(buffer.getvalue(), content_type='text/csv')
        filename = f"{dashboard_type}_analytics_{timezone.now().strftime('%Y%m%d')}.csv"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        return response

    def _generate_pdf(self, data, dashboard_type):
        """Generate PDF report."""
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import letter
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet
        except ImportError:
            return Response(
                {'error': 'reportlab not installed. Install with: pip install reportlab'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        styles = getSampleStyleSheet()
        elements = []

        # Title
        elements.append(Paragraph(f"{dashboard_type.title()} Analytics Report", styles['Title']))
        elements.append(Paragraph(f"Generated: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal']))
        elements.append(Spacer(1, 20))

        # Convert data to tables
        for key, value in data.items():
            if isinstance(value, dict):
                elements.append(Paragraph(key.replace('_', ' ').title(), styles['Heading2']))
                table_data = [[k.replace('_', ' ').title(), str(v) if v is not None else 'N/A'] for k, v in value.items()]
                if table_data:
                    t = Table(table_data)
                    t.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, 0), 10),
                        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                        ('GRID', (0, 0), (-1, -1), 1, colors.black)
                    ]))
                    elements.append(t)
                    elements.append(Spacer(1, 15))

        doc.build(elements)
        buffer.seek(0)

        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        filename = f"{dashboard_type}_analytics_{timezone.now().strftime('%Y%m%d')}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        return response


class RefreshCacheView(BaseAnalyticsAPIView):
    """
    API endpoint to refresh dashboard cache.

    POST /api/analytics/refresh-cache/

    Request Body:
    - dashboard_type: recruitment|diversity|hr|executive|all (default: all)

    Clears and regenerates cached dashboard data.
    """

    def post(self, request):
        dashboard_type = request.data.get('dashboard_type', 'all')

        service = DashboardDataService(use_cache=False)
        service.refresh_cache(dashboard_type)

        return Response({
            'status': 'success',
            'message': f'Cache refreshed for: {dashboard_type}',
            'refreshed_at': timezone.now().isoformat(),
        })


# ==================== LIST VIEWS FOR METRIC MODELS ====================

class RecruitmentMetricListView(generics.ListAPIView):
    """List recruitment metrics."""
    permission_classes = [IsAuthenticated]
    serializer_class = RecruitmentMetricSerializer
    queryset = RecruitmentMetric.objects.all()

    def get_queryset(self):
        queryset = super().get_queryset()
        period_type = self.request.query_params.get('period_type')
        if period_type:
            queryset = queryset.filter(period_type=period_type)
        return queryset


class DiversityMetricListView(generics.ListAPIView):
    """List diversity metrics."""
    permission_classes = [IsAuthenticated]
    serializer_class = DiversityMetricSerializer
    queryset = DiversityMetric.objects.all()

    def get_queryset(self):
        queryset = super().get_queryset()
        scope = self.request.query_params.get('scope')
        if scope:
            queryset = queryset.filter(scope=scope)
        return queryset


class HiringFunnelMetricListView(generics.ListAPIView):
    """List hiring funnel metrics."""
    permission_classes = [IsAuthenticated]
    serializer_class = HiringFunnelMetricSerializer
    queryset = HiringFunnelMetric.objects.all()


class TimeToHireMetricListView(generics.ListAPIView):
    """List time-to-hire metrics."""
    permission_classes = [IsAuthenticated]
    serializer_class = TimeToHireMetricSerializer
    queryset = TimeToHireMetric.objects.all()


class SourceEffectivenessMetricListView(generics.ListAPIView):
    """List source effectiveness metrics."""
    permission_classes = [IsAuthenticated]
    serializer_class = SourceEffectivenessMetricSerializer
    queryset = SourceEffectivenessMetric.objects.all()


class EmployeeRetentionMetricListView(generics.ListAPIView):
    """List employee retention metrics."""
    permission_classes = [IsAuthenticated]
    serializer_class = EmployeeRetentionMetricSerializer
    queryset = EmployeeRetentionMetric.objects.all()


class TimeOffAnalyticsListView(generics.ListAPIView):
    """List time-off analytics."""
    permission_classes = [IsAuthenticated]
    serializer_class = TimeOffAnalyticsSerializer
    queryset = TimeOffAnalytics.objects.all()


class PerformanceDistributionListView(generics.ListAPIView):
    """List performance distributions."""
    permission_classes = [IsAuthenticated]
    serializer_class = PerformanceDistributionSerializer
    queryset = PerformanceDistribution.objects.all()


# ==================== CYCLE 7 - ENHANCED DASHBOARD VIEWS ====================

from django.core.cache import cache
from .serializers import (
    DashboardMetricDisplaySerializer, TimeToHireDetailSerializer,
    SourceEffectivenessDetailSerializer, RecruitingFunnelDetailSerializer,
    TrendDataSerializer, ReportSerializer, ReportListSerializer,
    ReportGenerationRequestSerializer, ReportExportResultSerializer,
    DashboardSummarySerializer
)
import uuid as uuid_lib


class DashboardView(APIView):
    """
    Main analytics dashboard with all key metrics.

    GET /api/analytics/dashboard/

    Returns aggregated dashboard data with caching.
    """
    permission_classes = [IsAuthenticated]

    CACHE_TIMEOUT = 300  # 5 minutes

    def get(self, request):
        """
        Get all dashboard metrics.

        Query params:
        - start_date: Filter start date
        - end_date: Filter end date
        - period: day|week|month|quarter|year
        - refresh: Force cache refresh (default false)
        """
        date_filter = self._get_date_filter(request)
        refresh = request.query_params.get('refresh', 'false').lower() == 'true'

        # Cache key based on date range
        cache_key = f"dashboard_{date_filter.start_date}_{date_filter.end_date}"

        if not refresh:
            cached = cache.get(cache_key)
            if cached:
                return Response(cached)

        try:
            service = DashboardDataService(date_filter)

            # Build dashboard metrics
            metrics = self._build_metrics(service)
            charts = self._build_charts(service, date_filter)

            response_data = {
                'metrics': metrics,
                'charts': charts,
                'recent_activity': self._get_recent_activity(),
                'alerts': self._get_alerts(),
                'period': {
                    'start': date_filter.start_date.isoformat(),
                    'end': date_filter.end_date.isoformat(),
                    'period_type': date_filter.period,
                },
                'last_updated': timezone.now().isoformat(),
            }

            # Cache result
            cache.set(cache_key, response_data, self.CACHE_TIMEOUT)

            return Response(response_data)

        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    def _get_date_filter(self, request):
        """Parse date filter from request."""
        serializer = DateRangeSerializer(data=request.query_params)
        serializer.is_valid(raise_exception=True)
        data = serializer.validated_data

        return DateRangeFilter(
            start_date=data.get('start_date'),
            end_date=data.get('end_date'),
            period=data.get('period', 'month')
        )

    def _build_metrics(self, service):
        """Build dashboard metric cards."""
        recruitment = service.get_recruitment_dashboard()

        metrics = []

        # Open Positions
        open_positions = recruitment.get('job_metrics', {}).get('open_positions', 0)
        metrics.append({
            'metric_name': 'Open Positions',
            'metric_key': 'open_positions',
            'value': open_positions,
            'formatted_value': str(open_positions),
            'icon': 'briefcase',
            'color': '#4CAF50',
        })

        # Total Applications
        total_apps = recruitment.get('application_metrics', {}).get('total', 0)
        metrics.append({
            'metric_name': 'Total Applications',
            'metric_key': 'total_applications',
            'value': total_apps,
            'formatted_value': str(total_apps),
            'icon': 'users',
            'color': '#2196F3',
        })

        # Time to Hire
        time_to_hire = recruitment.get('time_to_hire', {}).get('avg_days', 0)
        metrics.append({
            'metric_name': 'Avg Time to Hire',
            'metric_key': 'avg_time_to_hire',
            'value': time_to_hire,
            'formatted_value': f"{time_to_hire:.1f} days" if time_to_hire else 'N/A',
            'unit': 'days',
            'icon': 'clock',
            'color': '#FF9800',
        })

        # Offer Acceptance Rate
        offer_rate = recruitment.get('conversion_rates', {}).get('offer_acceptance', 0)
        metrics.append({
            'metric_name': 'Offer Acceptance Rate',
            'metric_key': 'offer_acceptance_rate',
            'value': offer_rate,
            'formatted_value': f"{offer_rate:.1f}%" if offer_rate else 'N/A',
            'unit': '%',
            'icon': 'check-circle',
            'color': '#9C27B0',
        })

        return metrics

    def _build_charts(self, service, date_filter):
        """Build chart data for dashboard."""
        return {
            'applications_trend': {
                'labels': [],
                'datasets': [{
                    'label': 'Applications',
                    'data': [],
                    'borderColor': '#2196F3',
                    'fill': False,
                }],
            },
            'source_breakdown': {
                'labels': ['Career Page', 'LinkedIn', 'Referral', 'Other'],
                'datasets': [{
                    'data': [30, 25, 20, 25],
                    'backgroundColor': ['#4CAF50', '#2196F3', '#FF9800', '#9E9E9E'],
                }],
            },
        }

    def _get_recent_activity(self):
        """Get recent activity feed."""
        actions = UserAction.objects.select_related('user').order_by('-timestamp')[:10]
        return [
            {
                'id': action.id,
                'user': action.user.email if action.user else 'System',
                'action': action.get_action_type_display(),
                'description': action.description,
                'timestamp': action.timestamp.isoformat(),
            }
            for action in actions
        ]

    def _get_alerts(self):
        """Get system alerts."""
        alerts = []

        # Check for stale data
        stale_metrics = RecruitmentMetric.objects.filter(
            updated_at__lt=timezone.now() - timedelta(days=1)
        ).count()
        if stale_metrics > 0:
            alerts.append({
                'type': 'warning',
                'title': 'Stale Data',
                'message': f'{stale_metrics} metrics need refresh',
            })

        return alerts


class TimeToHireView(APIView):
    """
    Time to hire analytics endpoint.

    GET /api/analytics/time-to-hire/
    """
    permission_classes = [IsAuthenticated]

    CACHE_TIMEOUT = 600  # 10 minutes

    def get(self, request):
        """
        Get detailed time to hire metrics.

        Query params:
        - start_date, end_date, period
        - department: Filter by department
        - job_type: Filter by job type
        """
        date_filter = self._get_date_filter(request)
        department = request.query_params.get('department')
        job_type = request.query_params.get('job_type')

        cache_key = f"time_to_hire_{date_filter.start_date}_{date_filter.end_date}_{department}_{job_type}"
        cached = cache.get(cache_key)
        if cached:
            return Response(cached)

        try:
            # Get metrics from database
            filters = {
                'period_start__gte': date_filter.start_date,
                'period_end__lte': date_filter.end_date,
            }
            if department:
                filters['department'] = department
            if job_type:
                filters['job_type'] = job_type

            metrics = TimeToHireMetric.objects.filter(**filters).order_by('-period_start').first()

            if not metrics:
                # Generate response with defaults
                response_data = {
                    'average_days': 0,
                    'median_days': 0,
                    'min_days': None,
                    'max_days': None,
                    'by_department': {},
                    'by_source': {},
                    'trend': [],
                    'stage_breakdown': {},
                }
            else:
                response_data = {
                    'average_days': float(metrics.avg_time_to_hire or 0),
                    'median_days': float(metrics.median_time_to_hire or 0),
                    'min_days': int(metrics.min_time_to_hire) if metrics.min_time_to_hire else None,
                    'max_days': int(metrics.max_time_to_hire) if metrics.max_time_to_hire else None,
                    'by_department': metrics.by_source or {},
                    'by_source': metrics.by_source or {},
                    'trend': [],
                    'stage_breakdown': {
                        'screening': float(metrics.avg_time_in_screening or 0),
                        'interview': float(metrics.avg_time_in_interview or 0),
                        'assessment': float(metrics.avg_time_in_assessment or 0),
                        'offer': float(metrics.avg_time_in_offer or 0),
                    },
                    'target_days': metrics.target_time_to_fill,
                    'positions_within_target': metrics.positions_within_target,
                    'target_achievement_rate': float(metrics.target_achievement_rate or 0),
                }

            response_data['period'] = {
                'start': date_filter.start_date.isoformat(),
                'end': date_filter.end_date.isoformat(),
            }

            cache.set(cache_key, response_data, self.CACHE_TIMEOUT)
            return Response(response_data)

        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    def _get_date_filter(self, request):
        serializer = DateRangeSerializer(data=request.query_params)
        serializer.is_valid(raise_exception=True)
        data = serializer.validated_data
        return DateRangeFilter(
            start_date=data.get('start_date'),
            end_date=data.get('end_date'),
            period=data.get('period', 'month')
        )


class SourceAnalyticsView(APIView):
    """
    Source effectiveness analytics.

    GET /api/analytics/sources/
    """
    permission_classes = [IsAuthenticated]

    CACHE_TIMEOUT = 600

    def get(self, request):
        """
        Get source effectiveness metrics.

        Query params:
        - start_date, end_date
        - sort_by: effectiveness_score|hires|cost_per_hire
        """
        date_filter = self._get_date_filter(request)
        sort_by = request.query_params.get('sort_by', 'effectiveness_score')

        cache_key = f"source_analytics_{date_filter.start_date}_{date_filter.end_date}"
        cached = cache.get(cache_key)
        if cached:
            return Response(cached)

        try:
            metrics = SourceEffectivenessMetric.objects.filter(
                period_start__gte=date_filter.start_date,
                period_end__lte=date_filter.end_date,
            ).order_by(f'-{sort_by}' if sort_by != 'cost_per_hire' else sort_by)

            sources = []
            for m in metrics:
                sources.append({
                    'source': m.get_source_display(),
                    'source_key': m.source,
                    'applicants': m.total_applicants,
                    'qualified': m.qualified_applicants,
                    'interviewed': m.interviewed,
                    'hires': m.hires,
                    'hire_rate': float(m.hire_rate or 0),
                    'cost_per_hire': float(m.cost_per_hire or 0),
                    'effectiveness_score': float(m.effectiveness_score or 0),
                    'avg_performance_rating': float(m.avg_performance_rating or 0),
                    'retention_6_months': float(m.retention_rate_6_months or 0),
                })

            # Identify best/worst performers
            sorted_by_effectiveness = sorted(sources, key=lambda x: x['effectiveness_score'], reverse=True)
            best = sorted_by_effectiveness[:3] if len(sorted_by_effectiveness) >= 3 else sorted_by_effectiveness
            worst = sorted_by_effectiveness[-3:] if len(sorted_by_effectiveness) >= 3 else []

            # Calculate ROI by source
            roi_by_source = {}
            for s in sources:
                if s['cost_per_hire'] > 0:
                    roi_by_source[s['source_key']] = {
                        'roi_score': s['effectiveness_score'] / (s['cost_per_hire'] / 1000),
                        'cost_per_hire': s['cost_per_hire'],
                    }

            response_data = {
                'sources': sources,
                'best_performers': best,
                'worst_performers': worst,
                'roi_by_source': roi_by_source,
                'recommendations': self._generate_recommendations(sources),
                'period': {
                    'start': date_filter.start_date.isoformat(),
                    'end': date_filter.end_date.isoformat(),
                },
            }

            cache.set(cache_key, response_data, self.CACHE_TIMEOUT)
            return Response(response_data)

        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    def _get_date_filter(self, request):
        serializer = DateRangeSerializer(data=request.query_params)
        serializer.is_valid(raise_exception=True)
        data = serializer.validated_data
        return DateRangeFilter(
            start_date=data.get('start_date'),
            end_date=data.get('end_date'),
            period=data.get('period', 'month')
        )

    def _generate_recommendations(self, sources):
        """Generate recommendations based on source data."""
        recommendations = []

        # Find sources with high cost but low effectiveness
        for s in sources:
            if s['cost_per_hire'] > 5000 and s['effectiveness_score'] < 50:
                recommendations.append(
                    f"Consider reducing investment in {s['source']} due to low ROI"
                )
            if s['hire_rate'] > 10 and s['applicants'] < 20:
                recommendations.append(
                    f"Increase volume from {s['source']} - high quality but low volume"
                )

        return recommendations


class FunnelAnalyticsView(APIView):
    """
    Recruiting funnel analytics.

    GET /api/analytics/funnel/
    """
    permission_classes = [IsAuthenticated]

    CACHE_TIMEOUT = 600

    def get(self, request):
        """
        Get recruiting funnel data.

        Query params:
        - start_date, end_date
        - job_id: Filter by specific job
        - department: Filter by department
        """
        date_filter = self._get_date_filter(request)
        job_id = request.query_params.get('job_id')
        department = request.query_params.get('department')

        cache_key = f"funnel_{date_filter.start_date}_{date_filter.end_date}_{job_id}_{department}"
        cached = cache.get(cache_key)
        if cached:
            return Response(cached)

        try:
            filters = {
                'period_start__gte': date_filter.start_date,
                'period_end__lte': date_filter.end_date,
            }
            if job_id:
                filters['job_id'] = job_id
            if department:
                filters['department'] = department

            metric = HiringFunnelMetric.objects.filter(**filters).order_by('-period_start').first()

            if not metric:
                # Return empty funnel
                stages = [
                    {'stage': 'Applied', 'count': 0, 'conversion_rate': 100},
                    {'stage': 'Screening', 'count': 0, 'conversion_rate': 0},
                    {'stage': 'Interview', 'count': 0, 'conversion_rate': 0},
                    {'stage': 'Offer', 'count': 0, 'conversion_rate': 0},
                    {'stage': 'Hired', 'count': 0, 'conversion_rate': 0},
                ]
                response_data = {
                    'stages': stages,
                    'conversion_rates': {},
                    'bottlenecks': [],
                }
            else:
                stages = [
                    {
                        'stage': 'Applied',
                        'count': metric.stage_applied,
                        'conversion_rate': 100,
                        'drop_off': 0,
                    },
                    {
                        'stage': 'Screening',
                        'count': metric.stage_screening,
                        'conversion_rate': float(metric.rate_applied_to_screening or 0),
                        'drop_off': metric.dropped_at_screening,
                    },
                    {
                        'stage': 'Phone Interview',
                        'count': metric.stage_phone_interview,
                        'conversion_rate': float(metric.rate_screening_to_phone or 0),
                        'drop_off': metric.dropped_at_phone,
                    },
                    {
                        'stage': 'Technical',
                        'count': metric.stage_technical_assessment,
                        'conversion_rate': float(metric.rate_phone_to_technical or 0),
                        'drop_off': metric.dropped_at_technical,
                    },
                    {
                        'stage': 'Onsite',
                        'count': metric.stage_onsite_interview,
                        'conversion_rate': float(metric.rate_technical_to_onsite or 0),
                        'drop_off': metric.dropped_at_onsite,
                    },
                    {
                        'stage': 'Offer',
                        'count': metric.stage_offer,
                        'conversion_rate': float(metric.rate_reference_to_offer or 0),
                        'drop_off': metric.dropped_at_offer,
                    },
                    {
                        'stage': 'Hired',
                        'count': metric.stage_hired,
                        'conversion_rate': float(metric.rate_offer_to_hired or 0),
                        'drop_off': 0,
                    },
                ]

                # Identify bottlenecks (stages with lowest conversion)
                bottlenecks = []
                if metric.bottleneck_stage:
                    bottlenecks.append({
                        'stage': metric.bottleneck_stage,
                        'conversion_rate': float(metric.bottleneck_rate or 0),
                        'recommendation': 'Focus on improving this stage',
                    })

                response_data = {
                    'stages': stages,
                    'conversion_rates': {
                        'applied_to_screening': float(metric.rate_applied_to_screening or 0),
                        'screening_to_interview': float(metric.rate_screening_to_phone or 0),
                        'interview_to_offer': float(metric.rate_reference_to_offer or 0),
                        'offer_to_hired': float(metric.rate_offer_to_hired or 0),
                        'overall': float(metric.overall_conversion_rate or 0),
                    },
                    'bottlenecks': bottlenecks,
                    'average_time_in_funnel': float(metric.avg_days_in_funnel or 0),
                }

            response_data['period'] = {
                'start': date_filter.start_date.isoformat(),
                'end': date_filter.end_date.isoformat(),
            }

            cache.set(cache_key, response_data, self.CACHE_TIMEOUT)
            return Response(response_data)

        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    def _get_date_filter(self, request):
        serializer = DateRangeSerializer(data=request.query_params)
        serializer.is_valid(raise_exception=True)
        data = serializer.validated_data
        return DateRangeFilter(
            start_date=data.get('start_date'),
            end_date=data.get('end_date'),
            period=data.get('period', 'month')
        )


class TrendAnalyticsView(APIView):
    """
    Trend analysis for various metrics.

    GET /api/analytics/trends/
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        """
        Get trend data for charts.

        Query params:
        - metric: applications|hires|time_to_hire|offers
        - start_date, end_date
        - granularity: day|week|month
        """
        metric = request.query_params.get('metric', 'applications')
        granularity = request.query_params.get('granularity', 'month')
        date_filter = self._get_date_filter(request)

        try:
            # Generate trend data based on metric type
            if metric == 'applications':
                trend_data = self._get_applications_trend(date_filter, granularity)
            elif metric == 'hires':
                trend_data = self._get_hires_trend(date_filter, granularity)
            elif metric == 'time_to_hire':
                trend_data = self._get_time_to_hire_trend(date_filter, granularity)
            else:
                trend_data = {'labels': [], 'datasets': []}

            return Response({
                'labels': trend_data['labels'],
                'datasets': trend_data['datasets'],
                'period_type': granularity,
                'metric': metric,
                'period': {
                    'start': date_filter.start_date.isoformat(),
                    'end': date_filter.end_date.isoformat(),
                },
            })

        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    def _get_date_filter(self, request):
        serializer = DateRangeSerializer(data=request.query_params)
        serializer.is_valid(raise_exception=True)
        data = serializer.validated_data
        return DateRangeFilter(
            start_date=data.get('start_date'),
            end_date=data.get('end_date'),
            period=data.get('period', 'month')
        )

    def _get_applications_trend(self, date_filter, granularity):
        """Get applications trend data."""
        metrics = RecruitmentMetric.objects.filter(
            period_start__gte=date_filter.start_date,
            period_end__lte=date_filter.end_date,
        ).order_by('period_start')

        labels = []
        data = []
        for m in metrics:
            labels.append(m.period_start.strftime('%Y-%m-%d'))
            data.append(m.total_applications)

        return {
            'labels': labels,
            'datasets': [{
                'label': 'Applications',
                'data': data,
                'borderColor': '#2196F3',
                'backgroundColor': 'rgba(33, 150, 243, 0.1)',
                'fill': True,
            }],
        }

    def _get_hires_trend(self, date_filter, granularity):
        """Get hires trend data."""
        metrics = RecruitmentMetric.objects.filter(
            period_start__gte=date_filter.start_date,
            period_end__lte=date_filter.end_date,
        ).order_by('period_start')

        labels = []
        data = []
        for m in metrics:
            labels.append(m.period_start.strftime('%Y-%m-%d'))
            data.append(m.total_hires)

        return {
            'labels': labels,
            'datasets': [{
                'label': 'Hires',
                'data': data,
                'borderColor': '#4CAF50',
                'backgroundColor': 'rgba(76, 175, 80, 0.1)',
                'fill': True,
            }],
        }

    def _get_time_to_hire_trend(self, date_filter, granularity):
        """Get time to hire trend data."""
        metrics = TimeToHireMetric.objects.filter(
            period_start__gte=date_filter.start_date,
            period_end__lte=date_filter.end_date,
        ).order_by('period_start')

        labels = []
        data = []
        for m in metrics:
            labels.append(m.period_start.strftime('%Y-%m-%d'))
            data.append(float(m.avg_time_to_hire or 0))

        return {
            'labels': labels,
            'datasets': [{
                'label': 'Avg Days to Hire',
                'data': data,
                'borderColor': '#FF9800',
                'backgroundColor': 'rgba(255, 152, 0, 0.1)',
                'fill': True,
            }],
        }


class ReportsView(APIView):
    """
    Report management endpoint.

    GET /api/analytics/reports/ - List available reports
    POST /api/analytics/reports/ - Generate a new report
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        """
        List available reports.

        Query params:
        - report_type: Filter by type
        - page, page_size: Pagination
        """
        report_type = request.query_params.get('report_type')

        # For now, return predefined report types
        reports = [
            {
                'id': str(uuid_lib.uuid4()),
                'name': 'Recruitment Overview',
                'description': 'Complete recruitment metrics and KPIs',
                'report_type': 'recruitment',
                'created_by': 'System',
                'created_at': timezone.now().isoformat(),
            },
            {
                'id': str(uuid_lib.uuid4()),
                'name': 'Time to Hire Analysis',
                'description': 'Detailed time-to-hire metrics by department and source',
                'report_type': 'time_to_hire',
                'created_by': 'System',
                'created_at': timezone.now().isoformat(),
            },
            {
                'id': str(uuid_lib.uuid4()),
                'name': 'Source Effectiveness',
                'description': 'ROI analysis for candidate sources',
                'report_type': 'source_effectiveness',
                'created_by': 'System',
                'created_at': timezone.now().isoformat(),
            },
            {
                'id': str(uuid_lib.uuid4()),
                'name': 'Diversity Report',
                'description': 'Anonymized diversity metrics',
                'report_type': 'diversity',
                'created_by': 'System',
                'created_at': timezone.now().isoformat(),
            },
            {
                'id': str(uuid_lib.uuid4()),
                'name': 'Executive Summary',
                'description': 'High-level metrics for leadership',
                'report_type': 'executive',
                'created_by': 'System',
                'created_at': timezone.now().isoformat(),
            },
        ]

        if report_type:
            reports = [r for r in reports if r['report_type'] == report_type]

        return Response({
            'reports': reports,
            'total_count': len(reports),
        })

    def post(self, request):
        """
        Generate a new report.
        """
        serializer = ReportGenerationRequestSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        data = serializer.validated_data

        try:
            # Generate report based on type
            report_type = data['report_type']
            export_format = data.get('format', 'pdf')

            # Get date range
            start_date = data.get('start_date', timezone.now().date() - timedelta(days=30))
            end_date = data.get('end_date', timezone.now().date())
            date_filter = DateRangeFilter(start_date, end_date)

            service = DashboardDataService(date_filter)

            # Get data based on report type
            if report_type == 'recruitment':
                report_data = service.get_recruitment_dashboard()
            elif report_type == 'diversity':
                report_data = service.get_diversity_dashboard()
            elif report_type == 'hr':
                report_data = service.get_hr_dashboard()
            elif report_type == 'executive':
                report_data = service.get_executive_summary()
            else:
                report_data = service.get_recruitment_dashboard()

            # For now, return JSON format
            report_id = str(uuid_lib.uuid4())

            return Response({
                'report_id': report_id,
                'file_name': f'{report_type}_report_{timezone.now().strftime("%Y%m%d")}.{export_format}',
                'file_size': 0,
                'format': export_format,
                'generated_at': timezone.now().isoformat(),
                'expires_at': (timezone.now() + timedelta(hours=24)).isoformat(),
                'status': 'completed',
                'data': report_data if export_format == 'json' else None,
            })

        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class ReportExportView(APIView):
    """
    Export a specific report.

    GET /api/analytics/reports/<uuid:report_id>/export/
    """
    permission_classes = [IsAuthenticated]

    def get(self, request, report_id):
        """
        Export report to PDF/Excel.

        Query params:
        - format: pdf|excel|csv
        """
        export_format = request.query_params.get('format', 'pdf')

        try:
            # Generate export
            # For now, delegate to ExportReportView
            export_view = ExportReportView()
            request.data = {
                'format': export_format,
                'dashboard_type': 'all',
            }
            return export_view.post(request)

        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

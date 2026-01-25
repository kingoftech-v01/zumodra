"""
Analytics Services - Business Logic for Analytics Calculations

This module provides service classes for:
- RecruitmentAnalyticsService: Calculate recruitment KPIs
- DiversityAnalyticsService: EEOC-compliant diversity analytics
- HRAnalyticsService: HR metrics (retention, time-off, performance)
- DashboardDataService: Aggregate data for dashboards
"""

from decimal import Decimal
from datetime import date, datetime, timedelta
from typing import Optional, Dict, List, Any, Tuple
from collections import defaultdict
import statistics

from django.db import models
from django.db.models import (
    Count, Sum, Avg, Min, Max, Q, F,
    ExpressionWrapper, DurationField, DecimalField
)
from django.db.models.functions import (
    TruncDate, TruncWeek, TruncMonth, TruncQuarter, TruncYear,
    ExtractWeekDay, ExtractMonth, Coalesce
)
from django.utils import timezone
from django.core.cache import cache

from .models import (
    RecruitmentMetric, DiversityMetric, HiringFunnelMetric,
    TimeToHireMetric, SourceEffectivenessMetric, EmployeeRetentionMetric,
    TimeOffAnalytics, PerformanceDistribution, DashboardCache
)


class DateRangeFilter:
    """Helper class for handling date range filtering."""

    def __init__(
        self,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        period: str = 'month'
    ):
        self.end_date = end_date or timezone.now().date()

        if start_date:
            self.start_date = start_date
        else:
            # Default periods
            period_days = {
                'day': 1,
                'week': 7,
                'month': 30,
                'quarter': 90,
                'year': 365,
            }
            days = period_days.get(period, 30)
            self.start_date = self.end_date - timedelta(days=days)

        self.period = period

    def get_previous_period(self) -> Tuple[date, date]:
        """Get the equivalent previous period for comparison."""
        duration = self.end_date - self.start_date
        prev_end = self.start_date - timedelta(days=1)
        prev_start = prev_end - duration
        return prev_start, prev_end

    def get_date_range_filter(self, field_name: str = 'created_at') -> Q:
        """Return a Q object for filtering by date range."""
        return Q(**{
            f'{field_name}__gte': self.start_date,
            f'{field_name}__lte': self.end_date
        })


class RecruitmentAnalyticsService:
    """
    Service for calculating recruitment analytics and KPIs.
    """

    def __init__(self, date_filter: Optional[DateRangeFilter] = None):
        self.date_filter = date_filter or DateRangeFilter()
        # Lazy imports to avoid circular imports
        from jobs.models import JobPosting, Application, Interview, Offer, Candidate
        self.JobPosting = JobPosting
        self.Application = Application
        self.Interview = Interview
        self.Offer = Offer
        self.Candidate = Candidate

    def get_job_metrics(self) -> Dict[str, Any]:
        """Calculate job posting metrics."""
        jobs = self.JobPosting.objects.filter(
            self.date_filter.get_date_range_filter('created_at')
        )

        return {
            'total_jobs': jobs.count(),
            'open_positions': jobs.filter(status='open').count(),
            'new_positions': jobs.filter(
                published_at__gte=self.date_filter.start_date
            ).count(),
            'filled_positions': jobs.filter(status='filled').count(),
            'closed_positions': jobs.filter(
                status__in=['closed', 'cancelled']
            ).count(),
            'by_type': dict(jobs.values('job_type').annotate(
                count=Count('id')
            ).values_list('job_type', 'count')),
            'by_department': dict(jobs.values('category__name').annotate(
                count=Count('id')
            ).values_list('category__name', 'count')),
        }

    def get_application_metrics(self) -> Dict[str, Any]:
        """Calculate application metrics."""
        apps = self.Application.objects.filter(
            self.date_filter.get_date_range_filter('applied_at')
        )

        total = apps.count()

        status_counts = dict(apps.values('status').annotate(
            count=Count('id')
        ).values_list('status', 'count'))

        return {
            'total_applications': total,
            'new_applications': apps.filter(status='new').count(),
            'in_review': status_counts.get('in_review', 0),
            'shortlisted': status_counts.get('shortlisted', 0),
            'interviewing': status_counts.get('interviewing', 0),
            'rejected': status_counts.get('rejected', 0),
            'withdrawn': status_counts.get('withdrawn', 0),
            'hired': status_counts.get('hired', 0),
            'by_source': dict(apps.values('candidate__source').annotate(
                count=Count('id')
            ).values_list('candidate__source', 'count')),
        }

    def get_interview_metrics(self) -> Dict[str, Any]:
        """Calculate interview metrics."""
        interviews = self.Interview.objects.filter(
            self.date_filter.get_date_range_filter('scheduled_start')
        )

        return {
            'scheduled': interviews.count(),
            'completed': interviews.filter(status='completed').count(),
            'cancelled': interviews.filter(status='cancelled').count(),
            'no_show': interviews.filter(status='no_show').count(),
            'by_type': dict(interviews.values('interview_type').annotate(
                count=Count('id')
            ).values_list('interview_type', 'count')),
        }

    def get_offer_metrics(self) -> Dict[str, Any]:
        """Calculate offer metrics."""
        offers = self.Offer.objects.filter(
            self.date_filter.get_date_range_filter('created_at')
        )

        total = offers.count()
        accepted = offers.filter(status='accepted').count()
        declined = offers.filter(status='declined').count()

        acceptance_rate = None
        if total > 0:
            accepted_or_declined = accepted + declined
            if accepted_or_declined > 0:
                acceptance_rate = round((accepted / accepted_or_declined) * 100, 2)

        return {
            'total_offers': total,
            'accepted': accepted,
            'declined': declined,
            'pending': offers.filter(status='sent').count(),
            'expired': offers.filter(status='expired').count(),
            'acceptance_rate': acceptance_rate,
        }

    def get_conversion_rates(self) -> Dict[str, Any]:
        """Calculate funnel conversion rates."""
        apps = self.Application.objects.filter(
            self.date_filter.get_date_range_filter('applied_at')
        )

        total_applied = apps.count()
        total_interviewed = apps.filter(
            status__in=['interviewing', 'offer_pending', 'offer_extended', 'hired']
        ).count()
        total_offered = apps.filter(
            status__in=['offer_pending', 'offer_extended', 'hired']
        ).count()
        total_hired = apps.filter(status='hired').count()

        def safe_rate(num, denom):
            return round((num / denom) * 100, 2) if denom > 0 else None

        return {
            'applied_to_interview': safe_rate(total_interviewed, total_applied),
            'interview_to_offer': safe_rate(total_offered, total_interviewed),
            'offer_to_hire': safe_rate(total_hired, total_offered),
            'overall_conversion': safe_rate(total_hired, total_applied),
        }

    def get_time_to_hire_stats(self) -> Dict[str, Any]:
        """Calculate time-to-hire statistics."""
        hired_apps = self.Application.objects.filter(
            status='hired',
            hired_at__isnull=False
        ).filter(
            self.date_filter.get_date_range_filter('hired_at')
        )

        # Calculate days to hire for each application
        days_to_hire = []
        for app in hired_apps:
            if app.hired_at and app.applied_at:
                delta = app.hired_at - app.applied_at
                days_to_hire.append(delta.days)

        if not days_to_hire:
            return {
                'avg_days': None,
                'median_days': None,
                'min_days': None,
                'max_days': None,
                'total_hires': 0,
            }

        return {
            'avg_days': round(statistics.mean(days_to_hire), 2),
            'median_days': round(statistics.median(days_to_hire), 2),
            'min_days': min(days_to_hire),
            'max_days': max(days_to_hire),
            'total_hires': len(days_to_hire),
        }

    def calculate_recruitment_metric(self) -> RecruitmentMetric:
        """
        Calculate and create/update a RecruitmentMetric record.
        """
        job_metrics = self.get_job_metrics()
        app_metrics = self.get_application_metrics()
        interview_metrics = self.get_interview_metrics()
        offer_metrics = self.get_offer_metrics()
        conversion_rates = self.get_conversion_rates()
        time_stats = self.get_time_to_hire_stats()

        metric, created = RecruitmentMetric.objects.update_or_create(
            period_type='daily',
            period_start=self.date_filter.start_date,
            defaults={
                'period_end': self.date_filter.end_date,
                'open_positions': job_metrics['open_positions'],
                'new_positions': job_metrics['new_positions'],
                'filled_positions': job_metrics['filled_positions'],
                'closed_positions': job_metrics['closed_positions'],
                'total_applications': app_metrics['total_applications'],
                'new_applications': app_metrics['new_applications'],
                'applications_in_review': app_metrics['in_review'],
                'applications_shortlisted': app_metrics['shortlisted'],
                'applications_rejected': app_metrics['rejected'],
                'applications_withdrawn': app_metrics['withdrawn'],
                'interviews_scheduled': interview_metrics['scheduled'],
                'interviews_completed': interview_metrics['completed'],
                'interviews_cancelled': interview_metrics['cancelled'],
                'interviews_no_show': interview_metrics['no_show'],
                'offers_extended': offer_metrics['total_offers'],
                'offers_accepted': offer_metrics['accepted'],
                'offers_declined': offer_metrics['declined'],
                'offers_expired': offer_metrics['expired'],
                'total_hires': time_stats['total_hires'],
                'application_to_interview_rate': conversion_rates['applied_to_interview'],
                'interview_to_offer_rate': conversion_rates['interview_to_offer'],
                'offer_acceptance_rate': offer_metrics['acceptance_rate'],
                'overall_conversion_rate': conversion_rates['overall_conversion'],
                'avg_time_to_hire': time_stats['avg_days'],
                'by_department': job_metrics['by_department'],
                'by_job_type': job_metrics['by_type'],
            }
        )

        return metric

    def get_dashboard_data(self) -> Dict[str, Any]:
        """Get all recruitment data for dashboard display."""
        return {
            'job_metrics': self.get_job_metrics(),
            'application_metrics': self.get_application_metrics(),
            'interview_metrics': self.get_interview_metrics(),
            'offer_metrics': self.get_offer_metrics(),
            'conversion_rates': self.get_conversion_rates(),
            'time_to_hire': self.get_time_to_hire_stats(),
            'period': {
                'start': self.date_filter.start_date.isoformat(),
                'end': self.date_filter.end_date.isoformat(),
            },
        }


class DiversityAnalyticsService:
    """
    Service for calculating EEOC-compliant diversity analytics.
    Ensures anonymization by suppressing small category counts.
    """

    MIN_CATEGORY_SIZE = 5  # Minimum count to display (for privacy)

    def __init__(
        self,
        date_filter: Optional[DateRangeFilter] = None,
        scope: str = 'employees'
    ):
        self.date_filter = date_filter or DateRangeFilter()
        self.scope = scope
        # Lazy imports
        from hr_core.models import Employee
        from jobs.models import Candidate, Application
        self.Employee = Employee
        self.Candidate = Candidate
        self.Application = Application

    def _get_queryset(self):
        """Get the appropriate queryset based on scope."""
        if self.scope == 'employees':
            return self.Employee.objects.filter(
                status__in=['active', 'probation', 'on_leave']
            )
        elif self.scope == 'applicants':
            return self.Candidate.objects.filter(
                applications__applied_at__gte=self.date_filter.start_date,
                applications__applied_at__lte=self.date_filter.end_date,
            ).distinct()
        elif self.scope == 'hired':
            return self.Candidate.objects.filter(
                applications__status='hired',
                applications__hired_at__gte=self.date_filter.start_date,
                applications__hired_at__lte=self.date_filter.end_date,
            ).distinct()
        return self.Employee.objects.none()

    def _suppress_small_counts(self, data: Dict[str, int]) -> Dict[str, Any]:
        """
        Suppress counts below minimum threshold for privacy.
        Returns 'suppressed' for values < MIN_CATEGORY_SIZE.
        """
        result = {}
        for key, count in data.items():
            if 0 < count < self.MIN_CATEGORY_SIZE:
                result[key] = 'suppressed'
            else:
                result[key] = count
        return result

    def _calculate_percentages(
        self,
        data: Dict[str, Any],
        total: int
    ) -> Dict[str, Any]:
        """Calculate percentages, handling suppressed values."""
        result = {}
        for key, value in data.items():
            if value == 'suppressed' or total == 0:
                result[key] = None
            else:
                result[key] = round((value / total) * 100, 2)
        return result

    def get_gender_distribution(self) -> Dict[str, Any]:
        """
        Get gender distribution (anonymized).
        Note: This requires demographic data collection which may vary by implementation.
        """
        # Placeholder - actual implementation depends on how demographic data is stored
        # This would typically query a demographic/EEO data table

        return {
            'counts': {
                'male': 0,
                'female': 0,
                'nonbinary': 0,
                'not_disclosed': 0,
            },
            'percentages': {},
            'total': 0,
        }

    def get_ethnicity_distribution(self) -> Dict[str, Any]:
        """
        Get ethnicity distribution (EEOC categories, anonymized).
        """
        # Placeholder - actual implementation depends on EEO data collection

        return {
            'counts': {
                'white': 0,
                'black': 0,
                'hispanic': 0,
                'asian': 0,
                'native_american': 0,
                'pacific_islander': 0,
                'two_or_more': 0,
                'not_disclosed': 0,
            },
            'percentages': {},
            'total': 0,
        }

    def get_age_distribution(self) -> Dict[str, Any]:
        """
        Get age distribution by ranges (anonymized).
        """
        queryset = self._get_queryset()

        # Age ranges based on date of birth
        today = timezone.now().date()

        # Note: This requires a date_of_birth field on Employee/Candidate
        # Placeholder implementation

        return {
            'counts': {
                'under_25': 0,
                '25_34': 0,
                '35_44': 0,
                '45_54': 0,
                '55_64': 0,
                '65_plus': 0,
                'not_disclosed': 0,
            },
            'percentages': {},
            'total': 0,
        }

    def get_department_breakdown(self) -> Dict[str, Dict[str, Any]]:
        """
        Get diversity metrics by department (with anonymization).
        Only shows departments with sufficient data for each category.
        """
        # Placeholder - would aggregate by department
        return {}

    def calculate_diversity_metric(self) -> DiversityMetric:
        """
        Calculate and create/update a DiversityMetric record.
        """
        gender = self.get_gender_distribution()
        ethnicity = self.get_ethnicity_distribution()
        age = self.get_age_distribution()

        total = gender['total']

        metric, created = DiversityMetric.objects.update_or_create(
            period_type='quarterly',
            period_start=self.date_filter.start_date,
            scope=self.scope,
            defaults={
                'period_end': self.date_filter.end_date,
                'total_count': total,
                'gender_male_count': gender['counts'].get('male', 0),
                'gender_female_count': gender['counts'].get('female', 0),
                'gender_nonbinary_count': gender['counts'].get('nonbinary', 0),
                'gender_not_disclosed_count': gender['counts'].get('not_disclosed', 0),
                'ethnicity_white_count': ethnicity['counts'].get('white', 0),
                'ethnicity_black_count': ethnicity['counts'].get('black', 0),
                'ethnicity_hispanic_count': ethnicity['counts'].get('hispanic', 0),
                'ethnicity_asian_count': ethnicity['counts'].get('asian', 0),
                'ethnicity_native_american_count': ethnicity['counts'].get('native_american', 0),
                'ethnicity_pacific_islander_count': ethnicity['counts'].get('pacific_islander', 0),
                'ethnicity_two_or_more_count': ethnicity['counts'].get('two_or_more', 0),
                'ethnicity_not_disclosed_count': ethnicity['counts'].get('not_disclosed', 0),
                'age_under_25_count': age['counts'].get('under_25', 0),
                'age_25_34_count': age['counts'].get('25_34', 0),
                'age_35_44_count': age['counts'].get('35_44', 0),
                'age_45_54_count': age['counts'].get('45_54', 0),
                'age_55_64_count': age['counts'].get('55_64', 0),
                'age_65_plus_count': age['counts'].get('65_plus', 0),
                'age_not_disclosed_count': age['counts'].get('not_disclosed', 0),
                'min_category_size': self.MIN_CATEGORY_SIZE,
            }
        )

        # Calculate percentages
        metric.calculate_percentages()
        metric.save()

        return metric

    def get_dashboard_data(self, anonymize: bool = True) -> Dict[str, Any]:
        """Get diversity data for dashboard (anonymized by default)."""
        gender = self.get_gender_distribution()
        ethnicity = self.get_ethnicity_distribution()
        age = self.get_age_distribution()

        if anonymize:
            gender['counts'] = self._suppress_small_counts(gender['counts'])
            ethnicity['counts'] = self._suppress_small_counts(ethnicity['counts'])
            age['counts'] = self._suppress_small_counts(age['counts'])

        return {
            'gender': gender,
            'ethnicity': ethnicity,
            'age': age,
            'by_department': self.get_department_breakdown(),
            'scope': self.scope,
            'period': {
                'start': self.date_filter.start_date.isoformat(),
                'end': self.date_filter.end_date.isoformat(),
            },
            'anonymization_threshold': self.MIN_CATEGORY_SIZE,
        }


class HRAnalyticsService:
    """
    Service for HR analytics including retention, time-off, and performance.
    """

    def __init__(self, date_filter: Optional[DateRangeFilter] = None):
        self.date_filter = date_filter or DateRangeFilter()
        # Lazy imports
        from hr_core.models import (
            Employee, TimeOffRequest, PerformanceReview, Offboarding
        )
        self.Employee = Employee
        self.TimeOffRequest = TimeOffRequest
        self.PerformanceReview = PerformanceReview
        self.Offboarding = Offboarding

    def get_headcount_metrics(self) -> Dict[str, Any]:
        """Calculate headcount metrics."""
        employees = self.Employee.objects.all()
        active_statuses = ['active', 'probation', 'on_leave']

        # Current headcount
        current = employees.filter(status__in=active_statuses).count()

        # New hires in period
        new_hires = employees.filter(
            hire_date__gte=self.date_filter.start_date,
            hire_date__lte=self.date_filter.end_date,
        ).count()

        # Departures in period
        departures = employees.filter(
            termination_date__gte=self.date_filter.start_date,
            termination_date__lte=self.date_filter.end_date,
        ).count()

        # By department
        by_department = dict(
            employees.filter(status__in=active_statuses)
            .values('department__name')
            .annotate(count=Count('id'))
            .values_list('department__name', 'count')
        )

        # By employment type
        by_type = dict(
            employees.filter(status__in=active_statuses)
            .values('employment_type')
            .annotate(count=Count('id'))
            .values_list('employment_type', 'count')
        )

        return {
            'current_headcount': current,
            'new_hires': new_hires,
            'departures': departures,
            'net_change': new_hires - departures,
            'by_department': by_department,
            'by_employment_type': by_type,
        }

    def get_retention_metrics(self) -> Dict[str, Any]:
        """Calculate employee retention and turnover metrics."""
        employees = self.Employee.objects.all()
        active_statuses = ['active', 'probation', 'on_leave']

        # Starting headcount (at period start)
        starting = employees.filter(
            Q(hire_date__lt=self.date_filter.start_date) &
            (Q(termination_date__isnull=True) | Q(termination_date__gte=self.date_filter.start_date))
        ).count()

        # Ending headcount
        ending = employees.filter(status__in=active_statuses).count()

        # Average headcount
        avg_headcount = (starting + ending) / 2 if (starting + ending) > 0 else 0

        # Departures
        departures_qs = employees.filter(
            termination_date__gte=self.date_filter.start_date,
            termination_date__lte=self.date_filter.end_date,
        )
        total_departures = departures_qs.count()

        # Voluntary vs involuntary (based on status)
        voluntary = departures_qs.filter(status='resigned').count()
        involuntary = departures_qs.filter(status='terminated').count()

        # Calculate rates
        turnover_rate = None
        if avg_headcount > 0:
            turnover_rate = round((total_departures / avg_headcount) * 100, 2)

        # New hire retention (90-day)
        new_hires = employees.filter(
            hire_date__gte=self.date_filter.start_date - timedelta(days=90),
            hire_date__lte=self.date_filter.end_date - timedelta(days=90),
        )
        new_hire_count = new_hires.count()
        retained_new_hires = new_hires.filter(
            Q(termination_date__isnull=True) | Q(termination_date__gt=F('hire_date') + timedelta(days=90))
        ).count()

        new_hire_retention = None
        if new_hire_count > 0:
            new_hire_retention = round((retained_new_hires / new_hire_count) * 100, 2)

        return {
            'starting_headcount': starting,
            'ending_headcount': ending,
            'average_headcount': round(avg_headcount, 2),
            'total_departures': total_departures,
            'voluntary_departures': voluntary,
            'involuntary_departures': involuntary,
            'turnover_rate': turnover_rate,
            'retention_rate': round(100 - turnover_rate, 2) if turnover_rate else None,
            'new_hire_retention_rate': new_hire_retention,
        }

    def get_time_off_metrics(self) -> Dict[str, Any]:
        """Calculate time-off analytics."""
        requests = self.TimeOffRequest.objects.filter(
            start_date__lte=self.date_filter.end_date,
            end_date__gte=self.date_filter.start_date,
        )

        # By status
        approved = requests.filter(status='approved')

        # Total days by type
        by_type = {}
        for request in approved:
            type_name = request.time_off_type.name if request.time_off_type else 'Other'
            if type_name not in by_type:
                by_type[type_name] = Decimal('0')
            by_type[type_name] += request.total_days

        # Approval metrics
        total_requests = requests.count()
        approved_count = requests.filter(status='approved').count()
        rejected_count = requests.filter(status='rejected').count()
        pending_count = requests.filter(status='pending').count()

        approval_rate = None
        if (approved_count + rejected_count) > 0:
            approval_rate = round(
                (approved_count / (approved_count + rejected_count)) * 100, 2
            )

        return {
            'total_requests': total_requests,
            'approved': approved_count,
            'rejected': rejected_count,
            'pending': pending_count,
            'approval_rate': approval_rate,
            'days_by_type': {k: float(v) for k, v in by_type.items()},
        }

    def get_performance_metrics(self) -> Dict[str, Any]:
        """Calculate performance review metrics."""
        reviews = self.PerformanceReview.objects.filter(
            review_period_end__gte=self.date_filter.start_date,
            review_period_end__lte=self.date_filter.end_date,
        )

        total = reviews.count()
        completed = reviews.filter(status='completed').count()

        # Rating distribution
        completed_reviews = reviews.filter(
            status='completed',
            overall_rating__isnull=False
        )

        rating_dist = {
            5: completed_reviews.filter(overall_rating=5).count(),
            4: completed_reviews.filter(overall_rating=4).count(),
            3: completed_reviews.filter(overall_rating=3).count(),
            2: completed_reviews.filter(overall_rating=2).count(),
            1: completed_reviews.filter(overall_rating=1).count(),
        }

        # Average rating
        avg_rating = completed_reviews.aggregate(
            avg=Avg('overall_rating')
        )['avg']

        # Recommendations
        promotions = completed_reviews.filter(promotion_recommended=True).count()
        pips = completed_reviews.filter(pip_recommended=True).count()

        return {
            'total_reviews': total,
            'completed': completed,
            'completion_rate': round((completed / total) * 100, 2) if total > 0 else None,
            'rating_distribution': rating_dist,
            'average_rating': round(avg_rating, 2) if avg_rating else None,
            'promotion_recommendations': promotions,
            'pip_recommendations': pips,
        }

    def calculate_retention_metric(self) -> EmployeeRetentionMetric:
        """Calculate and create/update EmployeeRetentionMetric."""
        retention = self.get_retention_metrics()

        metric, created = EmployeeRetentionMetric.objects.update_or_create(
            period_type='monthly',
            period_start=self.date_filter.start_date,
            defaults={
                'period_end': self.date_filter.end_date,
                'starting_headcount': retention['starting_headcount'],
                'ending_headcount': retention['ending_headcount'],
                'average_headcount': retention['average_headcount'],
                'total_departures': retention['total_departures'],
                'voluntary_departures': retention['voluntary_departures'],
                'involuntary_departures': retention['involuntary_departures'],
                'overall_turnover_rate': retention['turnover_rate'],
                'overall_retention_rate': retention['retention_rate'],
                'new_hire_retention_rate': retention['new_hire_retention_rate'],
            }
        )

        return metric

    def calculate_time_off_analytics(self) -> TimeOffAnalytics:
        """Calculate and create/update TimeOffAnalytics."""
        time_off = self.get_time_off_metrics()
        headcount = self.get_headcount_metrics()

        analytics, created = TimeOffAnalytics.objects.update_or_create(
            period_type='monthly',
            period_start=self.date_filter.start_date,
            defaults={
                'period_end': self.date_filter.end_date,
                'total_employees': headcount['current_headcount'],
                'pto_requests': time_off['total_requests'],
                'requests_approved': time_off['approved'],
                'requests_rejected': time_off['rejected'],
                'requests_pending': time_off['pending'],
                'approval_rate': time_off['approval_rate'],
                'by_leave_type': time_off['days_by_type'],
            }
        )

        return analytics

    def calculate_performance_distribution(self) -> PerformanceDistribution:
        """Calculate and create/update PerformanceDistribution."""
        perf = self.get_performance_metrics()

        distribution, created = PerformanceDistribution.objects.update_or_create(
            period_type='yearly',
            review_cycle='annual',
            period_start=self.date_filter.start_date,
            defaults={
                'period_end': self.date_filter.end_date,
                'total_employees_reviewed': perf['total_reviews'],
                'reviews_completed': perf['completed'],
                'completion_rate': perf['completion_rate'],
                'rating_5_count': perf['rating_distribution'][5],
                'rating_4_count': perf['rating_distribution'][4],
                'rating_3_count': perf['rating_distribution'][3],
                'rating_2_count': perf['rating_distribution'][2],
                'rating_1_count': perf['rating_distribution'][1],
                'average_rating': perf['average_rating'],
                'promotion_recommendations': perf['promotion_recommendations'],
                'pip_recommendations': perf['pip_recommendations'],
            }
        )

        distribution.calculate_percentages()
        distribution.save()

        return distribution

    def get_dashboard_data(self) -> Dict[str, Any]:
        """Get all HR data for dashboard display."""
        return {
            'headcount': self.get_headcount_metrics(),
            'retention': self.get_retention_metrics(),
            'time_off': self.get_time_off_metrics(),
            'performance': self.get_performance_metrics(),
            'period': {
                'start': self.date_filter.start_date.isoformat(),
                'end': self.date_filter.end_date.isoformat(),
            },
        }


class DashboardDataService:
    """
    Service for aggregating dashboard data with caching.
    """

    CACHE_TIMEOUT = 300  # 5 minutes

    def __init__(
        self,
        date_filter: Optional[DateRangeFilter] = None,
        use_cache: bool = True
    ):
        self.date_filter = date_filter or DateRangeFilter()
        self.use_cache = use_cache

    def _get_cache_key(self, dashboard_type: str) -> str:
        """Generate cache key for dashboard data."""
        return f"dashboard:{dashboard_type}:{self.date_filter.start_date}:{self.date_filter.end_date}"

    def get_recruitment_dashboard(self) -> Dict[str, Any]:
        """Get recruitment dashboard data."""
        cache_key = self._get_cache_key('recruitment')

        if self.use_cache:
            cached = cache.get(cache_key)
            if cached:
                return cached

        service = RecruitmentAnalyticsService(self.date_filter)
        data = service.get_dashboard_data()

        # Add comparison with previous period
        prev_start, prev_end = self.date_filter.get_previous_period()
        prev_filter = DateRangeFilter(prev_start, prev_end)
        prev_service = RecruitmentAnalyticsService(prev_filter)

        data['comparison'] = {
            'period': {
                'start': prev_start.isoformat(),
                'end': prev_end.isoformat(),
            },
            'job_metrics': prev_service.get_job_metrics(),
            'application_metrics': prev_service.get_application_metrics(),
        }

        if self.use_cache:
            cache.set(cache_key, data, self.CACHE_TIMEOUT)

        return data

    def get_diversity_dashboard(self, scope: str = 'employees') -> Dict[str, Any]:
        """Get diversity dashboard data (anonymized)."""
        cache_key = self._get_cache_key(f'diversity:{scope}')

        if self.use_cache:
            cached = cache.get(cache_key)
            if cached:
                return cached

        service = DiversityAnalyticsService(self.date_filter, scope)
        data = service.get_dashboard_data(anonymize=True)

        if self.use_cache:
            cache.set(cache_key, data, self.CACHE_TIMEOUT)

        return data

    def get_hr_dashboard(self) -> Dict[str, Any]:
        """Get HR dashboard data."""
        cache_key = self._get_cache_key('hr')

        if self.use_cache:
            cached = cache.get(cache_key)
            if cached:
                return cached

        service = HRAnalyticsService(self.date_filter)
        data = service.get_dashboard_data()

        # Add comparison with previous period
        prev_start, prev_end = self.date_filter.get_previous_period()
        prev_filter = DateRangeFilter(prev_start, prev_end)
        prev_service = HRAnalyticsService(prev_filter)

        data['comparison'] = {
            'period': {
                'start': prev_start.isoformat(),
                'end': prev_end.isoformat(),
            },
            'headcount': prev_service.get_headcount_metrics(),
            'retention': prev_service.get_retention_metrics(),
        }

        if self.use_cache:
            cache.set(cache_key, data, self.CACHE_TIMEOUT)

        return data

    def get_executive_summary(self) -> Dict[str, Any]:
        """Get executive summary with key metrics from all areas."""
        cache_key = self._get_cache_key('executive')

        if self.use_cache:
            cached = cache.get(cache_key)
            if cached:
                return cached

        recruitment_service = RecruitmentAnalyticsService(self.date_filter)
        hr_service = HRAnalyticsService(self.date_filter)

        recruitment_data = recruitment_service.get_dashboard_data()
        hr_data = hr_service.get_dashboard_data()

        data = {
            'summary': {
                'open_positions': recruitment_data['job_metrics']['open_positions'],
                'total_applications': recruitment_data['application_metrics']['total_applications'],
                'hires': recruitment_data['time_to_hire']['total_hires'],
                'avg_time_to_hire': recruitment_data['time_to_hire']['avg_days'],
                'offer_acceptance_rate': recruitment_data['offer_metrics']['acceptance_rate'],
                'current_headcount': hr_data['headcount']['current_headcount'],
                'turnover_rate': hr_data['retention']['turnover_rate'],
                'average_performance_rating': hr_data['performance']['average_rating'],
            },
            'charts': {
                'applications_trend': self._get_applications_trend(),
                'hiring_funnel': self._get_funnel_data(recruitment_data),
                'turnover_trend': self._get_turnover_trend(),
            },
            'period': {
                'start': self.date_filter.start_date.isoformat(),
                'end': self.date_filter.end_date.isoformat(),
            },
        }

        if self.use_cache:
            cache.set(cache_key, data, self.CACHE_TIMEOUT)

        return data

    def _get_applications_trend(self) -> List[Dict[str, Any]]:
        """Get daily/weekly application trend data for charts."""
        from jobs.models import Application

        apps = Application.objects.filter(
            applied_at__gte=self.date_filter.start_date,
            applied_at__lte=self.date_filter.end_date,
        ).annotate(
            date=TruncDate('applied_at')
        ).values('date').annotate(
            count=Count('id')
        ).order_by('date')

        return [
            {'date': item['date'].isoformat(), 'count': item['count']}
            for item in apps
        ]

    def _get_funnel_data(self, recruitment_data: Dict) -> List[Dict[str, Any]]:
        """Format funnel data for chart display."""
        app_metrics = recruitment_data['application_metrics']

        return [
            {'stage': 'Applied', 'count': app_metrics['total_applications']},
            {'stage': 'Screened', 'count': app_metrics['shortlisted']},
            {'stage': 'Interviewing', 'count': app_metrics['interviewing']},
            {'stage': 'Offered', 'count': recruitment_data['offer_metrics']['total_offers']},
            {'stage': 'Hired', 'count': app_metrics['hired']},
        ]

    def _get_turnover_trend(self) -> List[Dict[str, Any]]:
        """Get monthly turnover trend data."""
        from hr_core.models import Employee

        # Get monthly departure counts
        departures = Employee.objects.filter(
            termination_date__gte=self.date_filter.start_date,
            termination_date__lte=self.date_filter.end_date,
        ).annotate(
            month=TruncMonth('termination_date')
        ).values('month').annotate(
            count=Count('id')
        ).order_by('month')

        return [
            {'month': item['month'].isoformat(), 'departures': item['count']}
            for item in departures
        ]

    def refresh_cache(self, dashboard_type: str = 'all') -> None:
        """Refresh dashboard cache."""
        types_to_refresh = ['recruitment', 'diversity', 'hr', 'executive']

        if dashboard_type != 'all':
            types_to_refresh = [dashboard_type]

        for dtype in types_to_refresh:
            cache_key = self._get_cache_key(dtype)
            cache.delete(cache_key)

        # Re-populate cache
        if 'recruitment' in types_to_refresh:
            self.get_recruitment_dashboard()
        if 'diversity' in types_to_refresh:
            self.get_diversity_dashboard()
        if 'hr' in types_to_refresh:
            self.get_hr_dashboard()
        if 'executive' in types_to_refresh:
            self.get_executive_summary()

    def save_to_db_cache(self, dashboard_type: str, data: Dict) -> DashboardCache:
        """Save dashboard data to database cache."""
        expires_at = timezone.now() + timedelta(seconds=self.CACHE_TIMEOUT)

        cache_obj, created = DashboardCache.objects.update_or_create(
            dashboard_type=dashboard_type,
            defaults={
                'data': data,
                'filters_applied': {
                    'start_date': self.date_filter.start_date.isoformat(),
                    'end_date': self.date_filter.end_date.isoformat(),
                },
                'expires_at': expires_at,
                'is_stale': False,
            }
        )

        return cache_obj


# ============================================================================
# ENHANCED SERVICES FOR CYCLE 7
# ============================================================================

class AnalyticsService:
    """
    Core analytics computations for hiring metrics.

    Provides centralized analytics calculations that can be used
    across different dashboards and reports.
    """

    def __init__(self, tenant_id: int = None, date_filter: DateRangeFilter = None):
        self.tenant_id = tenant_id
        self.date_filter = date_filter or DateRangeFilter()

    def compute_time_to_hire(self, period: str = 'month') -> Dict[str, Any]:
        """
        Compute time-to-hire metrics.

        Args:
            period: Aggregation period ('day', 'week', 'month', 'quarter', 'year')

        Returns:
            Dict with time-to-hire metrics
        """
        from jobs.models import Application

        # Get hired applications in period
        hired = Application.objects.filter(
            status='hired',
            hired_at__gte=self.date_filter.start_date,
            hired_at__lte=self.date_filter.end_date,
        )

        if self.tenant_id:
            hired = hired.filter(job__tenant_id=self.tenant_id)

        # Calculate time-to-hire for each
        times = []
        for app in hired.select_related('job'):
            if app.hired_at and app.applied_at:
                days = (app.hired_at - app.applied_at).days
                times.append(days)

        if not times:
            return {
                'avg_days': None,
                'median_days': None,
                'min_days': None,
                'max_days': None,
                'total_hires': 0,
                'by_department': {},
                'by_source': {},
                'trend': [],
            }

        return {
            'avg_days': round(statistics.mean(times), 2),
            'median_days': round(statistics.median(times), 2),
            'min_days': min(times),
            'max_days': max(times),
            'std_dev': round(statistics.stdev(times), 2) if len(times) > 1 else 0,
            'total_hires': len(times),
            'percentile_25': round(self._percentile(times, 25), 2),
            'percentile_75': round(self._percentile(times, 75), 2),
            'percentile_90': round(self._percentile(times, 90), 2),
        }

    def _percentile(self, data: List[float], percentile: int) -> float:
        """Calculate percentile of data."""
        if not data:
            return 0
        sorted_data = sorted(data)
        k = (len(sorted_data) - 1) * percentile / 100
        f = int(k)
        c = k - f
        if f + 1 < len(sorted_data):
            return sorted_data[f] * (1 - c) + sorted_data[f + 1] * c
        return sorted_data[f]

    def compute_source_effectiveness(self, period: str = 'month') -> Dict[str, Any]:
        """
        Compute effectiveness metrics for each candidate source.

        Returns:
            Dict mapping source names to effectiveness metrics
        """
        from jobs.models import Application, Candidate

        applications = Application.objects.filter(
            applied_at__gte=self.date_filter.start_date,
            applied_at__lte=self.date_filter.end_date,
        ).select_related('candidate')

        if self.tenant_id:
            applications = applications.filter(job__tenant_id=self.tenant_id)

        # Group by source
        source_data = defaultdict(lambda: {
            'total': 0,
            'interviewed': 0,
            'offered': 0,
            'hired': 0,
            'times_to_hire': [],
        })

        for app in applications:
            source = app.candidate.source if hasattr(app.candidate, 'source') else 'Unknown'
            source_data[source]['total'] += 1

            if app.status in ['interviewing', 'offer_pending', 'offer_extended', 'hired']:
                source_data[source]['interviewed'] += 1
            if app.status in ['offer_pending', 'offer_extended', 'hired']:
                source_data[source]['offered'] += 1
            if app.status == 'hired':
                source_data[source]['hired'] += 1
                if app.hired_at and app.applied_at:
                    source_data[source]['times_to_hire'].append(
                        (app.hired_at - app.applied_at).days
                    )

        # Calculate metrics for each source
        results = {}
        for source, data in source_data.items():
            total = data['total']
            results[source] = {
                'total_applicants': total,
                'interview_rate': round(data['interviewed'] / total * 100, 2) if total > 0 else 0,
                'offer_rate': round(data['offered'] / total * 100, 2) if total > 0 else 0,
                'hire_rate': round(data['hired'] / total * 100, 2) if total > 0 else 0,
                'avg_time_to_hire': (
                    round(statistics.mean(data['times_to_hire']), 2)
                    if data['times_to_hire'] else None
                ),
                'hires': data['hired'],
            }

        return results

    def compute_pipeline_velocity(self, pipeline_id: int = None) -> Dict[str, Any]:
        """
        Compute pipeline velocity metrics.

        Pipeline velocity measures how quickly candidates move through stages.

        Returns:
            Dict with velocity metrics by stage
        """
        from jobs.models import Application, PipelineStageChange

        # Get stage changes in period
        changes = PipelineStageChange.objects.filter(
            changed_at__gte=self.date_filter.start_date,
            changed_at__lte=self.date_filter.end_date,
        )

        if pipeline_id:
            changes = changes.filter(application__job__pipeline_id=pipeline_id)
        if self.tenant_id:
            changes = changes.filter(application__job__tenant_id=self.tenant_id)

        # Calculate time in each stage
        stage_times = defaultdict(list)

        for change in changes.select_related('from_stage', 'to_stage'):
            if change.from_stage and change.time_in_stage:
                stage_times[change.from_stage.name].append(change.time_in_stage.days)

        results = {}
        for stage, times in stage_times.items():
            results[stage] = {
                'avg_days': round(statistics.mean(times), 2) if times else 0,
                'median_days': round(statistics.median(times), 2) if times else 0,
                'candidates_processed': len(times),
            }

        return results

    def compute_recruiter_performance(
        self,
        recruiter_id: int = None
    ) -> Dict[str, Any]:
        """
        Compute recruiter performance metrics.

        Args:
            recruiter_id: Optional specific recruiter ID

        Returns:
            Dict with performance metrics (or list if no recruiter_id)
        """
        from django.contrib.auth import get_user_model
        from jobs.models import Application

        User = get_user_model()

        applications = Application.objects.filter(
            applied_at__gte=self.date_filter.start_date,
            applied_at__lte=self.date_filter.end_date,
        )

        if self.tenant_id:
            applications = applications.filter(job__tenant_id=self.tenant_id)
        if recruiter_id:
            applications = applications.filter(recruiter_id=recruiter_id)

        # Group by recruiter
        recruiter_data = defaultdict(lambda: {
            'screened': 0,
            'interviewed': 0,
            'offered': 0,
            'hired': 0,
            'times_to_hire': [],
        })

        for app in applications.select_related('recruiter'):
            if not app.recruiter:
                continue
            recruiter_id = app.recruiter_id

            recruiter_data[recruiter_id]['screened'] += 1

            if app.status in ['interviewing', 'offer_pending', 'offer_extended', 'hired']:
                recruiter_data[recruiter_id]['interviewed'] += 1
            if app.status in ['offer_pending', 'offer_extended', 'hired']:
                recruiter_data[recruiter_id]['offered'] += 1
            if app.status == 'hired':
                recruiter_data[recruiter_id]['hired'] += 1
                if app.hired_at and app.applied_at:
                    recruiter_data[recruiter_id]['times_to_hire'].append(
                        (app.hired_at - app.applied_at).days
                    )

        results = []
        for rec_id, data in recruiter_data.items():
            try:
                recruiter = User.objects.get(id=rec_id)
                name = recruiter.get_full_name() or recruiter.email
            except User.DoesNotExist:
                name = f"Recruiter {rec_id}"

            results.append({
                'recruiter_id': rec_id,
                'name': name,
                'candidates_screened': data['screened'],
                'candidates_interviewed': data['interviewed'],
                'offers_made': data['offered'],
                'hires': data['hired'],
                'avg_time_to_hire': (
                    round(statistics.mean(data['times_to_hire']), 2)
                    if data['times_to_hire'] else None
                ),
                'conversion_rate': (
                    round(data['hired'] / data['screened'] * 100, 2)
                    if data['screened'] > 0 else 0
                ),
            })

        # Sort by hires descending
        results.sort(key=lambda x: x['hires'], reverse=True)

        return results

    def get_hiring_trends(self, months: int = 12) -> List[Dict[str, Any]]:
        """
        Get hiring trends over time.

        Args:
            months: Number of months to include

        Returns:
            List of monthly hiring data points
        """
        from jobs.models import Application
        from django.db.models.functions import TruncMonth

        end_date = timezone.now().date()
        start_date = end_date - timedelta(days=months * 30)

        applications = Application.objects.filter(
            status='hired',
            hired_at__gte=start_date,
            hired_at__lte=end_date,
        )

        if self.tenant_id:
            applications = applications.filter(job__tenant_id=self.tenant_id)

        monthly_data = applications.annotate(
            month=TruncMonth('hired_at')
        ).values('month').annotate(
            hires=Count('id')
        ).order_by('month')

        # Also get applications received
        apps_received = Application.objects.filter(
            applied_at__gte=start_date,
            applied_at__lte=end_date,
        )
        if self.tenant_id:
            apps_received = apps_received.filter(job__tenant_id=self.tenant_id)

        apps_monthly = apps_received.annotate(
            month=TruncMonth('applied_at')
        ).values('month').annotate(
            applications=Count('id')
        ).order_by('month')

        # Combine data
        apps_dict = {item['month']: item['applications'] for item in apps_monthly}

        result = []
        for item in monthly_data:
            month = item['month']
            result.append({
                'month': month.isoformat() if month else None,
                'hires': item['hires'],
                'applications': apps_dict.get(month, 0),
                'conversion_rate': (
                    round(item['hires'] / apps_dict.get(month, 1) * 100, 2)
                    if apps_dict.get(month) else 0
                ),
            })

        return result


class ReportingService:
    """
    Service for generating reports in various formats.
    """

    def __init__(self, tenant_id: int = None, date_filter: DateRangeFilter = None):
        self.tenant_id = tenant_id
        self.date_filter = date_filter or DateRangeFilter()
        self.analytics_service = AnalyticsService(tenant_id, date_filter)

    def generate_recruiting_report(self) -> Dict[str, Any]:
        """
        Generate comprehensive recruiting report.

        Returns:
            Dict with all recruiting metrics
        """
        recruitment_service = RecruitmentAnalyticsService(self.date_filter)

        return {
            'title': 'Recruiting Report',
            'period': {
                'start': self.date_filter.start_date.isoformat(),
                'end': self.date_filter.end_date.isoformat(),
            },
            'generated_at': timezone.now().isoformat(),
            'metrics': {
                'jobs': recruitment_service.get_job_metrics(),
                'applications': recruitment_service.get_application_metrics(),
                'interviews': recruitment_service.get_interview_metrics(),
                'offers': recruitment_service.get_offer_metrics(),
                'conversions': recruitment_service.get_conversion_rates(),
                'time_to_hire': self.analytics_service.compute_time_to_hire(),
                'sources': self.analytics_service.compute_source_effectiveness(),
            },
            'trends': self.analytics_service.get_hiring_trends(6),
        }

    def generate_dei_report(self) -> Dict[str, Any]:
        """
        Generate Diversity, Equity, and Inclusion (DEI) report.

        Returns:
            Dict with DEI metrics (anonymized)
        """
        diversity_service = DiversityAnalyticsService(self.date_filter)

        return {
            'title': 'Diversity, Equity & Inclusion Report',
            'period': {
                'start': self.date_filter.start_date.isoformat(),
                'end': self.date_filter.end_date.isoformat(),
            },
            'generated_at': timezone.now().isoformat(),
            'note': 'Data is anonymized. Categories with <5 individuals are suppressed.',
            'current_employees': diversity_service.get_dashboard_data(anonymize=True),
            'applicants': DiversityAnalyticsService(
                self.date_filter, 'applicants'
            ).get_dashboard_data(anonymize=True),
            'new_hires': DiversityAnalyticsService(
                self.date_filter, 'hired'
            ).get_dashboard_data(anonymize=True),
        }

    def generate_cost_analysis(self) -> Dict[str, Any]:
        """
        Generate cost analysis report.

        Returns:
            Dict with cost metrics
        """
        from jobs.models import Application

        hired = Application.objects.filter(
            status='hired',
            hired_at__gte=self.date_filter.start_date,
            hired_at__lte=self.date_filter.end_date,
        )
        if self.tenant_id:
            hired = hired.filter(job__tenant_id=self.tenant_id)

        total_hires = hired.count()

        # Placeholder for actual cost data
        # In real implementation, this would come from a costs model
        estimated_cost_per_hire = Decimal('5000.00')  # Default estimate

        return {
            'title': 'Recruiting Cost Analysis',
            'period': {
                'start': self.date_filter.start_date.isoformat(),
                'end': self.date_filter.end_date.isoformat(),
            },
            'generated_at': timezone.now().isoformat(),
            'metrics': {
                'total_hires': total_hires,
                'estimated_cost_per_hire': float(estimated_cost_per_hire),
                'total_estimated_cost': float(estimated_cost_per_hire * total_hires),
                'by_source': self._get_cost_by_source(),
                'by_department': self._get_cost_by_department(),
            },
        }

    def _get_cost_by_source(self) -> Dict[str, float]:
        """Get estimated cost by source."""
        # Placeholder - would use actual cost data
        return {}

    def _get_cost_by_department(self) -> Dict[str, float]:
        """Get estimated cost by department."""
        # Placeholder - would use actual cost data
        return {}

    def export_to_excel(self, report: Dict) -> bytes:
        """
        Export report to Excel format.

        Args:
            report: Report data dict

        Returns:
            Excel file as bytes
        """
        try:
            import openpyxl
            from io import BytesIO

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = report.get('title', 'Report')[:31]

            # Write header
            ws['A1'] = report.get('title', 'Report')
            ws['A2'] = f"Period: {report.get('period', {}).get('start', '')} to {report.get('period', {}).get('end', '')}"
            ws['A3'] = f"Generated: {report.get('generated_at', '')}"

            # Write metrics
            row = 5
            for section, data in report.get('metrics', {}).items():
                ws.cell(row=row, column=1, value=section.title())
                row += 1

                if isinstance(data, dict):
                    for key, value in data.items():
                        ws.cell(row=row, column=1, value=key)
                        ws.cell(row=row, column=2, value=str(value) if value else 'N/A')
                        row += 1

                row += 1  # Blank row between sections

            # Save to bytes
            output = BytesIO()
            wb.save(output)
            return output.getvalue()

        except ImportError:
            # See TODO-ANALYTICS-001 in analytics/TODO.md
            raise NotImplementedError("openpyxl not installed")

    def export_to_pdf(self, report: Dict) -> bytes:
        """
        Export report to PDF format.

        Args:
            report: Report data dict

        Returns:
            PDF file as bytes
        """
        try:
            from reportlab.lib.pagesizes import letter
            from reportlab.pdfgen import canvas
            from io import BytesIO

            output = BytesIO()
            c = canvas.Canvas(output, pagesize=letter)
            width, height = letter

            # Title
            c.setFont("Helvetica-Bold", 16)
            c.drawString(72, height - 72, report.get('title', 'Report'))

            # Period
            c.setFont("Helvetica", 10)
            c.drawString(
                72, height - 100,
                f"Period: {report.get('period', {}).get('start', '')} to {report.get('period', {}).get('end', '')}"
            )

            # Metrics
            y_position = height - 140
            c.setFont("Helvetica", 10)

            for section, data in report.get('metrics', {}).items():
                if y_position < 100:
                    c.showPage()
                    y_position = height - 72

                c.setFont("Helvetica-Bold", 12)
                c.drawString(72, y_position, section.title())
                y_position -= 20

                c.setFont("Helvetica", 10)
                if isinstance(data, dict):
                    for key, value in data.items():
                        c.drawString(90, y_position, f"{key}: {value if value else 'N/A'}")
                        y_position -= 15

                y_position -= 10

            c.save()
            return output.getvalue()

        except ImportError:
            raise NotImplementedError("reportlab not installed")


class PredictiveAnalyticsService:
    """
    ML-powered predictive analytics for recruiting.

    Provides predictions for:
    - Time to fill positions
    - Offer acceptance probability
    - Employee retention risk
    - Hiring needs forecasting
    """

    def __init__(self, tenant_id: int = None):
        self.tenant_id = tenant_id

    def predict_time_to_fill(self, job) -> Dict[str, Any]:
        """
        Predict time to fill a job position.

        Uses historical data and job characteristics to estimate
        how long it will take to fill a position.

        Args:
            job: JobPosting instance

        Returns:
            Dict with prediction and confidence
        """
        from jobs.models import JobPosting

        # Get historical data for similar jobs
        similar_jobs = JobPosting.objects.filter(
            status='filled',
            filled_at__isnull=False,
        )

        if self.tenant_id:
            similar_jobs = similar_jobs.filter(tenant_id=self.tenant_id)

        # Filter by similar characteristics
        if hasattr(job, 'category') and job.category:
            similar_jobs = similar_jobs.filter(category=job.category)
        if hasattr(job, 'job_type') and job.job_type:
            similar_jobs = similar_jobs.filter(job_type=job.job_type)
        if hasattr(job, 'location') and job.location:
            similar_jobs = similar_jobs.filter(location__icontains=job.location[:10])

        # Calculate fill times
        fill_times = []
        for j in similar_jobs[:100]:  # Limit for performance
            if j.filled_at and j.published_at:
                days = (j.filled_at - j.published_at).days
                fill_times.append(days)

        if not fill_times:
            # No historical data - return industry average
            return {
                'predicted_days': 42,  # Industry average ~6 weeks
                'confidence': 'low',
                'confidence_score': 0.3,
                'range_min': 21,
                'range_max': 63,
                'factors': ['No historical data available'],
            }

        avg_days = statistics.mean(fill_times)
        std_dev = statistics.stdev(fill_times) if len(fill_times) > 1 else avg_days * 0.3

        # Confidence based on sample size
        if len(fill_times) >= 20:
            confidence = 'high'
            confidence_score = 0.8
        elif len(fill_times) >= 10:
            confidence = 'medium'
            confidence_score = 0.6
        else:
            confidence = 'low'
            confidence_score = 0.4

        return {
            'predicted_days': round(avg_days),
            'confidence': confidence,
            'confidence_score': confidence_score,
            'range_min': max(7, round(avg_days - std_dev)),
            'range_max': round(avg_days + std_dev),
            'sample_size': len(fill_times),
            'factors': self._identify_time_factors(job, fill_times),
        }

    def _identify_time_factors(self, job, historical_times: List[int]) -> List[str]:
        """Identify factors that may affect time to fill."""
        factors = []

        avg = statistics.mean(historical_times) if historical_times else 42

        # Job type factors
        if hasattr(job, 'job_type'):
            if job.job_type == 'contract':
                factors.append("Contract positions typically fill faster")
            elif job.job_type == 'executive':
                factors.append("Executive searches typically take longer")

        # Remote factor
        if hasattr(job, 'is_remote') and job.is_remote:
            factors.append("Remote positions may attract more candidates")

        # Urgency
        if hasattr(job, 'priority') and job.priority == 'high':
            factors.append("High priority may accelerate process")

        return factors

    def predict_offer_acceptance(self, candidate, offer) -> Dict[str, Any]:
        """
        Predict probability of offer acceptance.

        Args:
            candidate: Candidate instance
            offer: Offer instance

        Returns:
            Dict with acceptance probability and factors
        """
        from jobs.models import Offer

        # Get historical acceptance data
        past_offers = Offer.objects.filter(
            status__in=['accepted', 'declined'],
        )

        if self.tenant_id:
            past_offers = past_offers.filter(job__tenant_id=self.tenant_id)

        accepted = past_offers.filter(status='accepted').count()
        total = past_offers.count()

        if total == 0:
            base_rate = 0.75  # Industry average
        else:
            base_rate = accepted / total

        # Adjust based on factors
        probability = base_rate
        factors = []

        # Salary competitiveness
        if hasattr(offer, 'salary') and hasattr(candidate, 'expected_salary'):
            if offer.salary >= candidate.expected_salary:
                probability += 0.1
                factors.append("Offer meets salary expectations (+)")
            else:
                gap = (candidate.expected_salary - offer.salary) / candidate.expected_salary
                probability -= gap * 0.2
                factors.append(f"Offer below expectations ({gap:.0%} gap) (-)")

        # Response time factor
        if hasattr(offer, 'created_at') and hasattr(offer.application, 'applied_at'):
            days_to_offer = (offer.created_at.date() - offer.application.applied_at.date()).days
            if days_to_offer <= 14:
                probability += 0.05
                factors.append("Quick process (<14 days) (+)")
            elif days_to_offer > 30:
                probability -= 0.1
                factors.append("Long process (>30 days) (-)")

        probability = max(0.1, min(0.95, probability))

        # Confidence based on data
        confidence = 'high' if total >= 50 else ('medium' if total >= 20 else 'low')

        return {
            'probability': round(probability, 2),
            'confidence': confidence,
            'factors': factors,
            'recommendation': (
                'Strong accept' if probability >= 0.8 else
                'Likely accept' if probability >= 0.6 else
                'Uncertain' if probability >= 0.4 else
                'At risk'
            ),
        }

    def predict_employee_retention(self, employee) -> Dict[str, Any]:
        """
        Predict employee retention risk.

        Args:
            employee: Employee instance

        Returns:
            Dict with retention prediction
        """
        from hr_core.models import Employee

        # Get historical retention data
        departed = Employee.objects.filter(
            status__in=['resigned', 'terminated'],
            termination_date__isnull=False,
        )

        if self.tenant_id:
            departed = departed.filter(tenant_id=self.tenant_id)

        # Calculate average tenure of departed employees
        tenures = []
        for emp in departed[:200]:
            if emp.hire_date and emp.termination_date:
                years = (emp.termination_date - emp.hire_date).days / 365
                tenures.append(years)

        avg_tenure = statistics.mean(tenures) if tenures else 3.0  # Default 3 years

        # Calculate employee's current tenure
        if hasattr(employee, 'hire_date') and employee.hire_date:
            current_tenure = (timezone.now().date() - employee.hire_date).days / 365
        else:
            current_tenure = 0

        # Base risk calculation
        # Risk increases after average tenure
        if current_tenure < avg_tenure * 0.5:
            base_risk = 0.2  # Early stage
        elif current_tenure < avg_tenure:
            base_risk = 0.3  # Building up
        elif current_tenure < avg_tenure * 1.5:
            base_risk = 0.4  # Peak risk
        else:
            base_risk = 0.3  # Likely to stay

        risk_factors = []
        retention_risk = base_risk

        # Performance factor
        if hasattr(employee, 'latest_performance_rating'):
            rating = employee.latest_performance_rating
            if rating and rating <= 2:
                retention_risk += 0.2
                risk_factors.append("Low performance rating (+risk)")
            elif rating and rating >= 4:
                retention_risk -= 0.1
                risk_factors.append("High performance (-risk)")

        # Promotion factor
        if hasattr(employee, 'last_promotion_date'):
            if employee.last_promotion_date:
                years_since = (timezone.now().date() - employee.last_promotion_date).days / 365
                if years_since > 3:
                    retention_risk += 0.15
                    risk_factors.append(f"No promotion in {years_since:.1f} years (+risk)")

        retention_risk = max(0.05, min(0.95, retention_risk))

        return {
            'retention_risk': round(retention_risk, 2),
            'risk_level': (
                'High' if retention_risk >= 0.6 else
                'Medium' if retention_risk >= 0.4 else
                'Low'
            ),
            'current_tenure_years': round(current_tenure, 1),
            'avg_tenure_departed': round(avg_tenure, 1),
            'risk_factors': risk_factors,
            'recommendation': (
                'Immediate attention needed' if retention_risk >= 0.6 else
                'Monitor closely' if retention_risk >= 0.4 else
                'Routine engagement'
            ),
        }

    def forecast_hiring_needs(
        self,
        department_id: int = None,
        months: int = 6
    ) -> List[Dict[str, Any]]:
        """
        Forecast hiring needs based on historical patterns.

        Args:
            department_id: Optional department filter
            months: Months to forecast

        Returns:
            List of monthly forecasts
        """
        from jobs.models import JobPosting
        from hr_core.models import Employee
        from django.db.models.functions import TruncMonth

        # Get historical hiring patterns
        past_months = 12
        start_date = timezone.now().date() - timedelta(days=past_months * 30)

        jobs_query = JobPosting.objects.filter(
            status='filled',
            filled_at__gte=start_date,
        )
        if self.tenant_id:
            jobs_query = jobs_query.filter(tenant_id=self.tenant_id)
        if department_id:
            jobs_query = jobs_query.filter(category_id=department_id)

        monthly_hires = list(jobs_query.annotate(
            month=TruncMonth('filled_at')
        ).values('month').annotate(
            count=Count('id')
        ).order_by('month').values_list('count', flat=True))

        # Get turnover rate
        employees = Employee.objects.filter(status__in=['active', 'probation'])
        if self.tenant_id:
            employees = employees.filter(tenant_id=self.tenant_id)
        if department_id:
            employees = employees.filter(department_id=department_id)

        headcount = employees.count()
        monthly_turnover_rate = 0.01  # Default 1% monthly

        # Calculate average hiring
        avg_monthly_hires = statistics.mean(monthly_hires) if monthly_hires else 2

        # Generate forecast
        forecasts = []
        current_date = timezone.now().date()

        for i in range(months):
            forecast_date = current_date + timedelta(days=30 * (i + 1))

            # Forecast = historical average + turnover replacement
            turnover_replacement = headcount * monthly_turnover_rate
            forecast_hires = avg_monthly_hires + turnover_replacement

            # Add seasonal adjustment (placeholder)
            month = forecast_date.month
            seasonal_factor = 1.0
            if month in [1, 6, 9]:  # Common hiring seasons
                seasonal_factor = 1.2
            elif month in [12]:  # Holiday slowdown
                seasonal_factor = 0.7

            forecast_hires *= seasonal_factor

            forecasts.append({
                'month': forecast_date.replace(day=1).isoformat(),
                'forecast_hires': round(forecast_hires),
                'turnover_replacement': round(turnover_replacement, 1),
                'growth_hires': round(forecast_hires - turnover_replacement),
                'confidence': 'medium' if len(monthly_hires) >= 6 else 'low',
            })

        return forecasts
